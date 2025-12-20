from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable
from zipfile import ZipFile

import pytest
from attrs import define, field
from openpyxl import Workbook

from exdrf_xl.column import XlColumn
from exdrf_xl.schema import XlSchema
from exdrf_xl.table import XlTable


@define(slots=True, kw_only=True)
class _Col(XlColumn["_Table", dict[str, Any]]):
    key: str

    def value_from_record(self, record: dict[str, Any]) -> Any:
        return record.get(self.key)


@define(slots=True, kw_only=True)
class _Table(XlTable[dict[str, Any]]):
    records: list[dict[str, Any]] = field(factory=list, repr=False)

    def get_selector(self):  # pragma: no cover
        raise NotImplementedError

    def get_rows(self, session) -> Iterable[dict[str, Any]]:
        yield from self.records


class TestXlTableFormatting:
    def test_applies_widths_and_alignments_and_table_range(self):
        schema = XlSchema()
        table = _Table(
            schema=schema,
            sheet_name="Sheet1",
            xl_name="T1",
            records=[
                {"a": "x", "b": "y", "id": 1},
                {"a": "x2", "b": "y2", "id": 1},
            ],
            columns=[
                _Col(
                    xl_name="A",
                    key="a",
                    col_width=33.0,
                    wrap_text=True,
                    h_align="center",
                    v_align="top",
                    font_color="FF0000",
                    bg_color="00FF00",
                ),
                _Col(
                    xl_name="B",
                    key="b",
                    col_width=11.0,
                    wrap_text=False,
                    h_align="right",
                    v_align="bottom",
                    number_format="0.00",
                ),
                _Col(
                    xl_name="id",
                    key="id",
                    col_width=10.0,
                    wrap_text=False,
                    h_align="left",
                    v_align="center",
                ),
            ],
        )

        wb = Workbook()
        ws = wb.active

        # The implementation ignores the session object for this test table.
        table.write_to_sheet(ws, session=None)  # type: ignore[arg-type]

        # Column access via `table[...]` should work for both index and name.
        assert table[0].xl_name == "A"
        assert table["id"].xl_name == "id"

        assert ws["A1"].value == "A"
        assert ws["B1"].value == "B"
        assert ws["C1"].value == "id"

        # Data must start at row 2.
        assert ws["A2"].value == "x"
        assert ws["B2"].value == "y"
        assert ws["C2"].value == 1
        assert ws["C3"].value == 1

        # Number format should be applied only when explicitly configured.
        assert ws["B2"].number_format == "0.00"

        # The Excel structured table must cover the data row.
        assert "T1" in ws.tables
        assert ws.tables["T1"].ref == "A1:C3"

        # Widths must be applied from column definitions.
        assert ws.column_dimensions["A"].width == pytest.approx(33.0)
        assert ws.column_dimensions["B"].width == pytest.approx(11.0)
        assert ws.column_dimensions["C"].width == pytest.approx(10.0)

        # Alignments must be applied from column definitions (data rows).
        a2_align = ws["A2"].alignment
        assert a2_align.wrap_text is True
        assert a2_align.horizontal == "center"
        assert a2_align.vertical == "top"

        b2_align = ws["B2"].alignment
        assert b2_align.wrap_text is False
        assert b2_align.horizontal == "right"
        assert b2_align.vertical == "bottom"

        # Styles should only be applied when explicitly configured.
        assert ws["A2"].font.color is not None
        assert ws["A2"].font.color.rgb == "FFFF0000"
        assert ws["A2"].fill.patternType == "solid"
        assert ws["A2"].fill.start_color.rgb == "FF00FF00"

        # Duplicate IDs should be highlighted via conditional formatting.
        assert len(ws.conditional_formatting) == 1
        rules = ws.conditional_formatting["C2:C3"]
        assert len(rules) == 1
        rule = rules[0]
        assert rule.type == "duplicateValues"
        assert rule.formula == ()

        # Verify the saved XML matches Excel's duplicateValues form (no formula).
        buff = BytesIO()
        wb.save(buff)
        buff.seek(0)
        with ZipFile(buff) as zf:
            sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
        assert 'conditionalFormatting sqref="C2:C3"' in sheet_xml
        assert 'cfRule type="duplicateValues"' in sheet_xml
        assert "<formula>" not in sheet_xml
