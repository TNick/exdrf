import logging
from typing import TYPE_CHECKING, cast

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableColumn
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def read_column_widths_from_existing_file(
    wb: "Workbook",
) -> dict[str, dict[str, float]]:
    """Read column widths from existing Excel file.

    Args:
        wb: The workbook to read column widths from.

    Returns:
        Dictionary mapping table names to dictionaries of column names to
        widths.
    """
    column_widths: dict[str, dict[str, float]] = {}

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]

        if not isinstance(ws, Worksheet):
            continue

        # For each Table object in the worksheet's tables
        for table_name, table in ws.tables.items():
            table_widths: dict[str, float] = {}
            table = cast("Table", table)
            if not table.ref:
                logger.error("Missing ref in table %s", table_name)
                continue

            # Get table boundaries to determine column range
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if (
                min_col is None
                or min_row is None
                or max_col is None
                or max_row is None
            ):
                logger.error(
                    "Invalid table range: "
                    "min_col %s, "
                    "min_row %s, "
                    "max_col %s, "
                    "max_row %s",
                    min_col,
                    min_row,
                    max_col,
                    max_row,
                )
                continue

            # Get column names from tableColumns
            if not table.tableColumns:
                logger.debug("Missing tableColumns")
                continue

            # Map each table column to its width
            for col_idx, table_col in enumerate(table.tableColumns, start=0):
                table_col = cast("TableColumn", table_col)
                col_name = table_col.name
                assert col_name

                # Calculate the actual column index in the worksheet
                ws_col_idx = min_col + col_idx
                if ws_col_idx > max_col:
                    break

                # Get column letter and width
                col_letter = get_column_letter(ws_col_idx)
                col_dim = ws.column_dimensions.get(col_letter)

                if col_dim:
                    if col_dim.width:
                        table_widths[col_name] = col_dim.width
                    else:
                        logger.debug("Width missing in column %s", col_letter)
                else:
                    logger.debug("Dimension missing in column %s", col_letter)

            column_widths[table_name] = table_widths

    return column_widths
