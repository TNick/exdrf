"""Utilities for creating Excel hyperlinks between structured tables."""

from __future__ import annotations

from copy import copy
from typing import TYPE_CHECKING, Any

from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from exdrf_xl.schema import XlSchema
    from exdrf_xl.table import XlTable


def xl_link_column_to_table_key(
    *,
    ws: "Worksheet",
    schema: "XlSchema",
    source_table: "XlTable",
    source_column: Any,
    row_count: int,
    target_table_name: str,
    target_key_column_name: str,
    target_sheet_name: str | None = None,
    apply_hyperlink_style: bool = True,
) -> None:
    """Convert a source column into hyperlinks to a target table's key column.

    The link points to the target sheet cell in the key column on the matching
    row. Matching is done using Excel `MATCH()` against the structured reference
    `target_table[target_key_column]`.

    Notes:
    - Values are embedded as Excel constants in the hyperlink formula to avoid
      circular references when the formula is written back into the same cell.
    - Empty cells and cells that already contain a formula are left unchanged.

    Args:
        ws: Worksheet containing the source table.
        schema: Workbook schema that can resolve `target_table_name`.
        source_table: Table owning the source column.
        source_column: Column reference in `source_table` (either a column
            object instance or the column name).
        row_count: Number of data rows written in the table.
        target_table_name: Structured table name of the target table.
        target_key_column_name: Column name in the target table that contains
            the key values to match against.
        target_sheet_name: Optional explicit worksheet name. If not provided,
            it is derived from the target table (`sheet_name[0:31]`).
        apply_hyperlink_style: If True, apply a typical hyperlink style
            (blue + underline) while preserving other font properties.
    """

    # No data rows -> nothing to link.
    if row_count <= 0:
        return

    # Resolve the source column index in the sheet.
    src_col_idx_0 = source_table.get_column_index(source_column)
    if src_col_idx_0 < 0:
        return

    # Resolve the target table and its key column letter (in the target sheet).
    target_tbl = schema.get_table(target_table_name)
    if target_tbl is None:
        return

    target_key_col_idx_0 = target_tbl.get_column_index(target_key_column_name)
    if target_key_col_idx_0 < 0:
        return

    sheet_name = (
        target_sheet_name
        if target_sheet_name is not None
        else target_tbl.sheet_name[0:31]
    )
    sheet_name = sheet_name.replace("'", "''")
    target_key_col_letter = get_column_letter(target_key_col_idx_0 + 1)

    def _to_excel_constant(value: Any) -> str:
        """Convert a Python value into an Excel literal constant.

        Args:
            value: Python value read from a worksheet cell.

        Returns:
            An Excel literal suitable for embedding in a formula.
        """
        if value is None:
            return '""'
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, (int, float)):
            return str(value)
        escaped = str(value).replace('"', '""')
        return '"%s"' % escaped

    max_row = row_count + 1  # include header row 1; data starts at 2
    for row_idx in range(2, max_row + 1):
        cell: Any = ws.cell(row=row_idx, column=src_col_idx_0 + 1)

        raw = cell.value
        if raw is None:
            continue
        if isinstance(raw, str) and raw.strip() == "":
            continue
        if isinstance(raw, str) and raw.lstrip().startswith("="):
            continue

        key = _to_excel_constant(raw)
        link_base = "\"#'%s'!%s\"" % (sheet_name, target_key_col_letter)

        # Keep the displayed value identical to the cell's original value.
        cell.value = (
            "=IFERROR("
            "HYPERLINK("
            "%s & (MATCH(%s, %s[%s], 0) + 1),"
            "%s"
            "),"
            "%s"
            ")"
            % (
                link_base,
                key,
                target_table_name,
                target_key_column_name,
                key,
                key,
            )
        )

        if not apply_hyperlink_style:
            continue

        new_font = copy(cell.font)
        new_font.color = Color(rgb="FF0563C1")
        new_font.underline = "single"
        cell.font = new_font
