from __future__ import annotations

import logging
import os.path
from dataclasses import dataclass
from dataclasses import field as dataclass_field
from datetime import datetime
from typing import Any, Callable, TypeAlias

from attrs import define, field
from exdrf.constants import FIELD_TYPE_DT  # type: ignore[import]
from exdrf.field_types.date_time import UNKNOWN_DATETIME
from exdrf_util.rotate_backups import rotate_backups
from openpyxl import Workbook, load_workbook  # type: ignore[import]
from sqlalchemy import and_, or_, select
from sqlalchemy.exc import DataError, IntegrityError
from sqlalchemy.orm import Session

from exdrf_xl.utils.col_widths import read_column_widths_from_existing_file
from exdrf_xl.utils.toposort import toposort_tables

from .table import XlTable

logger = logging.getLogger(__name__)


XlRecord: TypeAlias = dict[str, Any]
DbRecord: TypeAlias = Any
OldRecord: TypeAlias = tuple[DbRecord, XlRecord]
TableChanges: TypeAlias = tuple["XlTable[Any]", list[XlRecord], list[OldRecord]]
SchemaChanges: TypeAlias = list[TableChanges]


def _default_is_db_pk(value: Any) -> bool:
    """Return True when a value represents a database primary key.

    The current heuristic is intentionally simple and customizable by callers:
    - integers are considered database IDs
    - strings are considered placeholders for new records

    Note:
        `bool` is a subclass of `int` in Python. We exclude booleans.

    Args:
        value: Candidate primary key value.

    Returns:
        True if `value` is considered a DB primary key.
    """
    return isinstance(value, int) and not isinstance(value, bool)


def _normalize_unknown_datetime(value: Any) -> Any:
    """Normalize unknown date-time sentinel to a canonical representation."""
    if isinstance(value, str) and value.strip().lower() == "x":
        return UNKNOWN_DATETIME
    if isinstance(value, datetime):
        if (
            value.year == 1000
            and value.month == 2
            and value.day == 3
            and value.hour == 4
            and value.minute == 5
            and value.second == 6
        ):
            return UNKNOWN_DATETIME
    return value


def _is_datetime_type(type_name: Any) -> bool:
    """Return True if the column `type_name` represents a datetime."""
    return str(type_name) in ("datetime", FIELD_TYPE_DT)


def _build_import_table_ref_map(
    tables: list["XlTable[Any]"],
) -> dict[str, str]:
    """Build a reference-name -> canonical table name mapping.

    Generated code may refer to other tables by `XlTable.xl_name` or by the
    `XlTable` subclass name (e.g. `Org`), depending on generator/template.
    This helper maps both to a canonical identifier (`XlTable.xl_name`) so the
    import logic can be robust.

    Args:
        tables: Tables included in the import plan.

    Returns:
        Mapping from reference names to canonical `XlTable.xl_name`.
    """
    ref_to_xl_name: dict[str, str] = {}
    for t in tables:
        ref_to_xl_name[t.xl_name] = t.xl_name
        ref_to_xl_name[t.__class__.__name__] = t.xl_name
    return ref_to_xl_name


def _iter_chunks(items: list[Any], size: int) -> list[list[Any]]:
    """Split a list into consecutive chunks."""
    if size <= 0:
        raise ValueError("Invalid chunk size %r" % size)
    return [items[i : i + size] for i in range(0, len(items), size)]


@dataclass(frozen=True)
class CellDiff:
    """A single-cell difference between the database and Excel."""

    column: str
    old_value: Any
    new_value: Any


@dataclass(frozen=True)
class RowDiff:
    """Differences for a single row in a table."""

    table_name: str
    is_new: bool
    pk: dict[str, Any]
    xl_row: XlRecord
    db_rec: DbRecord | None
    diffs: tuple[CellDiff, ...]


@dataclass(frozen=True)
class TableDiff:
    """Changes for a single table."""

    table: "XlTable[Any]"
    new_rows: tuple[RowDiff, ...]
    modified_rows: tuple[RowDiff, ...]
    existing_rows: int
    total_rows: int


@dataclass(frozen=True)
class ImportPlan:
    """A prepared import plan, including detected modifications."""

    source_path: str
    tables: tuple[TableDiff, ...]

    @property
    def has_changes(self) -> bool:
        return any(t.new_rows or t.modified_rows for t in self.tables)


@dataclass
class PendingUpdate:
    """A deferred placeholder replacement.

    Attributes:
        db_rec: The SQLAlchemy instance to update.
        column_name: Attribute name on the db record to set.
        placeholder: Placeholder token that must be resolved to an integer id.
        fk_table_name: Name of the table that the placeholder belongs to (the
            table referenced by the foreign key column).
    """

    db_rec: DbRecord | None
    column_name: str
    placeholder: str
    fk_table_name: str


@dataclass
class ApplyResult:
    """Result of applying an import plan.

    Attributes:
        inserted: Number of rows inserted.
        updated: Number of rows updated.
        deferred: Number of deferred foreign key updates applied.
        placeholder_to_id: Mapping from (table_name, placeholder_string)
            to allocated integer ID. Only includes placeholders that were
            resolved during import.
    """

    inserted: int = 0
    updated: int = 0
    deferred: int = 0
    placeholder_to_id: dict[tuple[str, str], int] = dataclass_field(
        default_factory=dict
    )


@define
class XlSchema:
    """Describes the mapping between a database schema and an
    Excel workbook.

    Attributes:
        tables: The tables in the schema, exported in order.
    """

    tables: list["XlTable"] = field(factory=list, repr=False)

    def export_to_file(self, db: Any, path: str, max_backups: int = 10):
        """Exports the content of the selected tables to an Excel file.

        Args:
            db: Database connection used to create a session for exporting.
            path: Output `.xlsx` file path.
        """
        widths_map = {}
        if os.path.exists(path):
            try:
                wb = load_workbook(path, data_only=True)
                widths_map = read_column_widths_from_existing_file(wb)
            except Exception:
                logger.error(
                    "Failed to retrieve information from previous version",
                    exc_info=True,
                )
            rotate_backups(path, max_backups=max_backups)

        wb = Workbook()
        with db.same_session() as session:
            self.before_export(wb, session)
            for table in self.tables:
                sheet = wb.create_sheet(title=table.sheet_name[0:31])
                table.write_to_sheet(
                    sheet, session, col_widths=widths_map.get(table.xl_name, {})
                )
            self.after_export(wb, session)
        wb.calculation.fullCalcOnLoad = True  # type: ignore
        wb.save(path)

    def plan_import_from_file(
        self,
        db: Any,
        path: str,
        *,
        is_db_pk: Callable[[Any], bool] | None = None,
        batch_size: int = 100,
        wb: Any | None = None,
    ) -> ImportPlan:
        """Build an import plan from an Excel workbook.

        This inspects each configured `XlTable` present in the workbook and
        classifies each row as:
        - new row (primary key looks like a placeholder), or
        - existing row (primary key looks like a DB id and can be found), or
        - modified existing row (existing row with changed values)

        Args:
            db: Database connection used to create a session for comparisons.
            path: Input `.xlsx` file path (used if `wb` is None).
            is_db_pk: Optional predicate used to determine whether a primary key
                cell comes from the database. Defaults to `_default_is_db_pk`.
            batch_size: Batch size for database queries.
            wb: Optional workbook to use. If provided, `path` is only used for
                reference. If None, the workbook is loaded from `path`.

        Returns:
            An `ImportPlan` describing all detected new/modified rows.
        """
        is_db_pk = is_db_pk or _default_is_db_pk

        should_close = False
        if wb is None:
            wb = load_workbook(filename=path, read_only=False, data_only=True)
            should_close = True

        assert wb is not None  # For type checker.

        try:
            with db.same_session() as session:
                table_plans: list[TableDiff] = []

                for table in self.tables:
                    ws_name = table.sheet_name[0:31]
                    if ws_name not in wb.sheetnames:
                        continue
                    ws = wb[ws_name]

                    ws_table = ws.tables.get(table.xl_name)
                    if ws_table is None:
                        continue

                    pk_cols = [c.xl_name for c in table.columns if c.primary]
                    new_rows: list[RowDiff] = []
                    modified_rows: list[RowDiff] = []
                    existing_rows = 0

                    xl_rows_all: list[XlRecord] = list(
                        table.iter_excel_table(ws, ws_table)
                    )
                    total_rows = len(xl_rows_all)

                    # Try the batched path if the table exposes a db model.
                    db_model = None
                    try:
                        db_model = table.get_db_model_class()
                    except Exception:
                        db_model = None

                    pk_cols_eff = (
                        table.get_pk_column_names()
                        if hasattr(table, "get_pk_column_names")
                        else pk_cols
                    )

                    def add_new_row(xl_rec: XlRecord) -> None:
                        pk = {
                            k: xl_rec.get(k) for k in pk_cols_eff if k in xl_rec
                        }
                        diffs = tuple(
                            CellDiff(
                                column=c.xl_name,
                                old_value=None,
                                new_value=xl_rec.get(c.xl_name, None),
                            )
                            for c in table.columns
                            if not bool(getattr(c, "read_only", False))
                            if c.xl_name in xl_rec
                        )
                        new_rows.append(
                            RowDiff(
                                table_name=table.xl_name,
                                is_new=True,
                                pk=pk,
                                xl_row=xl_rec,
                                db_rec=None,
                                diffs=diffs,
                            )
                        )

                    def add_existing_row(db_rec: Any, xl_rec: XlRecord) -> None:
                        nonlocal existing_rows
                        existing_rows += 1
                        pk = {
                            k: xl_rec.get(k) for k in pk_cols_eff if k in xl_rec
                        }
                        cell_diffs: list[CellDiff] = []
                        for c in table.columns:
                            if bool(getattr(c, "read_only", False)):
                                continue
                            if c.xl_name not in xl_rec:
                                continue
                            new_val = xl_rec.get(c.xl_name, None)
                            old_val = c.value_from_record(db_rec)

                            # Normalize unknown datetime sentinel so we don't
                            # report diffs due to Excel range limitations.
                            if _is_datetime_type(getattr(c, "type_name", None)):
                                new_val = _normalize_unknown_datetime(new_val)
                                old_val = _normalize_unknown_datetime(old_val)

                            if isinstance(new_val, str):
                                new_val = new_val.strip()
                            if isinstance(old_val, str):
                                old_val = old_val.strip()

                            if old_val != new_val:
                                cell_diffs.append(
                                    CellDiff(
                                        column=c.xl_name,
                                        old_value=old_val,
                                        new_value=new_val,
                                    )
                                )

                        if cell_diffs:
                            modified_rows.append(
                                RowDiff(
                                    table_name=table.xl_name,
                                    is_new=False,
                                    pk=pk,
                                    xl_row=xl_rec,
                                    db_rec=db_rec,
                                    diffs=tuple(cell_diffs),
                                )
                            )

                    # Batched lookup path (preferred).
                    if db_model is not None and pk_cols_eff:
                        pk_key_to_xl: dict[tuple[Any, ...], XlRecord] = {}
                        keyed_conditions: list[
                            tuple[tuple[Any, ...], tuple[Any, ...]]
                        ] = []

                        for xl_rec in xl_rows_all:
                            pk_key = tuple(
                                xl_rec.get(k, None) for k in pk_cols_eff
                            )

                            conds = None
                            try:
                                conds = table.pk_conditions(xl_rec, is_db_pk)
                            except Exception:
                                logger.exception(
                                    "PK conditions failed for table %s",
                                    table.xl_name,
                                )
                                conds = None

                            if conds is None:
                                add_new_row(xl_rec)
                                continue

                            pk_key_to_xl[pk_key] = xl_rec
                            keyed_conditions.append((pk_key, conds))

                        found_keys: set[tuple[Any, ...]] = set()
                        for chunk in _iter_chunks(keyed_conditions, batch_size):
                            try:
                                where_clause = or_(
                                    *[and_(*conds) for _, conds in chunk]
                                )
                                stmt = select(db_model).where(where_clause)
                                for db_rec in session.scalars(stmt):
                                    db_key = tuple(
                                        getattr(db_rec, k) for k in pk_cols_eff
                                    )
                                    xl_rec_opt = pk_key_to_xl.get(db_key)
                                    if xl_rec_opt is None:
                                        continue
                                    found_keys.add(db_key)
                                    add_existing_row(db_rec, xl_rec_opt)
                            except DataError:
                                # Postgres aborts the transaction on DataError;
                                # rollback before continuing with fallback.
                                session.rollback()
                                # Fallback to per-row lookup for this chunk.
                                for pk_key, _conds in chunk:
                                    xl_rec_opt = pk_key_to_xl.get(pk_key)
                                    if xl_rec_opt is None:
                                        continue
                                    db_rec = None
                                    try:
                                        db_rec = table.find_db_rec(
                                            session, xl_rec_opt, is_db_pk
                                        )
                                    except DataError:
                                        session.rollback()
                                        db_rec = None

                                    if db_rec is None:
                                        add_new_row(xl_rec_opt)
                                    else:
                                        found_keys.add(pk_key)
                                        add_existing_row(db_rec, xl_rec_opt)

                        # Any DB-looking PKs not found are treated as new.
                        for pk_key, _conds in keyed_conditions:
                            if pk_key in found_keys:
                                continue
                            xl_rec_opt = pk_key_to_xl.get(pk_key)
                            if xl_rec_opt is not None:
                                add_new_row(xl_rec_opt)

                    else:
                        # Legacy fallback: per-row `find_db_rec`.
                        for xl_rec in xl_rows_all:
                            pk_is_db = pk_cols_eff and all(
                                is_db_pk(xl_rec.get(k)) for k in pk_cols_eff
                            )

                            db_rec = None
                            if pk_is_db:
                                try:
                                    db_rec = table.find_db_rec(
                                        session, xl_rec, is_db_pk
                                    )
                                except DataError:
                                    session.rollback()
                                    db_rec = None

                            if db_rec is None:
                                add_new_row(xl_rec)
                            else:
                                add_existing_row(db_rec, xl_rec)

                    if new_rows or modified_rows:
                        table_plans.append(
                            TableDiff(
                                table=table,
                                new_rows=tuple(new_rows),
                                modified_rows=tuple(modified_rows),
                                existing_rows=existing_rows,
                                total_rows=total_rows,
                            )
                        )

                return ImportPlan(source_path=path, tables=tuple(table_plans))
        finally:
            if should_close:
                wb.close()

    def import_from_file(
        self,
        db: Any,
        path: str,
        *,
        accept_new: bool = True,
        accept_modified: bool = True,
        is_db_pk: Callable[[Any], bool] | None = None,
    ):
        """Import data from an Excel file into the database.

        This is a non-interactive apply step; review/selection should be done
        by calling `plan_import_from_file()` + `render_review_html()` first.

        Args:
            db: Database connection used to create a session for import.
            path: Input `.xlsx` file path.
            accept_new: Whether to insert new rows.
            accept_modified: Whether to apply modifications to existing rows.
            is_db_pk: Optional predicate used during planning. Defaults to
                `_default_is_db_pk`.
        """
        plan = self.plan_import_from_file(db, path, is_db_pk=is_db_pk)
        self.apply_import_plan(
            db,
            plan,
            accept_new=accept_new,
            accept_modified=accept_modified,
            is_db_pk=is_db_pk,
        )

    def apply_import_plan(
        self,
        db: Any,
        plan: ImportPlan,
        *,
        accept_new: bool,
        accept_modified: bool,
        is_db_pk: Callable[[Any], bool] | None = None,
    ) -> ApplyResult:
        """Apply a previously computed `ImportPlan` to the database.

        Placeholder handling:
        - New records can use string placeholders in the primary key column
          `id`. After insertion, the allocated integer id is recorded and any
          occurrences of that placeholder in other `*_id` columns can be
          replaced.
        - For unresolved placeholders in `*_id` columns, updates may be deferred
          until all inserts complete.

        This is designed to be flexible enough for future row-level selection
        UIs: callers can filter `plan.tables[*].new_rows/modified_rows` before
        passing them here.

        Args:
            db: Database connection used to create a session for import.
            plan: Import plan previously created by `plan_import_from_file()`.
            accept_new: Whether to insert new rows.
            accept_modified: Whether to apply modifications to existing rows.
            is_db_pk: Predicate used to determine DB IDs. Defaults to
                `_default_is_db_pk`.

        Returns:
            An `ApplyResult` with counts.
        """
        is_db_pk = is_db_pk or _default_is_db_pk

        # Build worklists.
        to_insert: list[tuple[XlTable[Any], RowDiff]] = []
        to_update: list[tuple[XlTable[Any], RowDiff]] = []
        for table_plan in plan.tables:
            if accept_new:
                to_insert.extend(
                    (table_plan.table, r) for r in table_plan.new_rows
                )
            if accept_modified:
                to_update.extend(
                    (table_plan.table, r) for r in table_plan.modified_rows
                )

        # Fast exit.
        if not to_insert and not to_update:
            return ApplyResult()

        # Determine best-effort table insertion order based on FK dependencies
        # between tables that are included in this import.
        import_tables = [table_plan.table for table_plan in plan.tables]
        ref_to_xl_name = _build_import_table_ref_map(import_tables)

        deps: dict[str, set[str]] = {
            tbl.xl_name: set() for tbl in import_tables
        }
        table_fk_cols: dict[str, set[str]] = {
            tbl.xl_name: set() for tbl in import_tables
        }
        for tbl in import_tables:
            for c in tbl.columns:
                fk_ref = c.fk_table
                if not fk_ref:
                    continue

                # Only consider FK relations to tables that are part of this
                # import set; otherwise the column should behave like a normal
                # data column (no placeholder mapping / deferral).
                dep = ref_to_xl_name.get(str(fk_ref))
                if not dep:
                    continue
                if dep == tbl.xl_name:
                    continue

                deps[tbl.xl_name].add(dep)
                table_fk_cols[tbl.xl_name].add(c.xl_name)

        table_names_in_plan_order = [tbl.xl_name for tbl in import_tables]
        insertion_order = toposort_tables(table_names_in_plan_order, deps)
        order_idx = {name: i for i, name in enumerate(insertion_order)}

        # Reorder insert worklist by table insertion order; keep row order
        # within each table stable.
        to_insert.sort(key=lambda x: order_idx.get(x[0].xl_name, 10**9))

        result = ApplyResult()

        # Placeholder -> db id mapping, scoped by table name.
        # Key is (table_name, placeholder_string), value is the integer id.
        placeholder_to_id: dict[tuple[str, str], int] = {}

        # Deferred placeholder updates for unresolved foreign keys.
        pending_updates: list[PendingUpdate] = []

        def _primary_cols(table: XlTable[Any]) -> list[str]:
            return [c.xl_name for c in table.columns if c.primary]

        def _is_generated_id_pk(table: XlTable[Any]) -> bool:
            pk = _primary_cols(table)
            return pk == ["id"]

        def _fk_cols_in_import_set(table: XlTable[Any]) -> set[str]:
            return table_fk_cols.get(table.xl_name, set())

        def _transform_xl_rec_for_apply(
            table: XlTable[Any],
            xl_rec: XlRecord,
            *,
            allow_defer_fk: bool,
        ) -> tuple[XlRecord, str | None, list[str], list[PendingUpdate]]:
            """Prepare an Excel record dict for applying to a db record.

            Returns:
                (transformed_record, id_placeholder, unresolved_pk_columns,
                pending_updates)
            """
            transformed = dict(xl_rec)
            pk_cols = _primary_cols(table)
            unresolved_pk: list[str] = []
            row_pending: list[PendingUpdate] = []
            fk_cols_in_plan = _fk_cols_in_import_set(table)

            # Handle placeholder primary key for auto-generated id.
            id_placeholder: str | None = None
            if _is_generated_id_pk(table):
                raw_id = transformed.get("id", None)
                if raw_id is not None and not is_db_pk(raw_id):
                    if isinstance(raw_id, str):
                        id_placeholder = raw_id.strip()
                    transformed["id"] = None

            # Resolve placeholders in likely FK columns.
            # Find column objects to access fk_table information.
            col_by_name: dict[str, Any] = {}
            for c in table.get_included_columns():
                col_by_name[c.xl_name] = c

            for key, value in list(transformed.items()):
                if value is None:
                    continue

                # Normalize empty strings to None only for FK columns that point
                # to tables included in this import run. For other tables (not
                # part of the import set), treat the value as normal user data.
                if (
                    key in fk_cols_in_plan
                    and key.endswith("_id")
                    and isinstance(value, str)
                ):
                    v = value.strip()
                    if v == "":
                        transformed[key] = None
                        continue
                    value = v
                    transformed[key] = v

                if not isinstance(value, str):
                    continue

                placeholder = value.strip()
                if placeholder == "":
                    continue

                # Only treat placeholders specially for FK columns that point to
                # tables included in this import run.
                if key not in fk_cols_in_plan:
                    continue

                # Get the foreign key table name for this column.
                col = col_by_name.get(key)
                fk_table_name = None
                if col is not None:
                    fk_table_name = getattr(col, "fk_table", None)
                if fk_table_name is None:
                    # Fallback: if we can't determine the FK table, skip
                    # placeholder resolution for this column.
                    continue

                # Look up placeholder scoped to the target table.
                placeholder_key = (fk_table_name, placeholder)
                mapped = placeholder_to_id.get(placeholder_key)
                if mapped is not None:
                    transformed[key] = mapped
                    continue

                # Unresolved placeholder in PK -> row can't be inserted yet.
                if key in pk_cols and key in fk_cols_in_plan:
                    unresolved_pk.append(key)
                    continue

                # Unresolved placeholder in fk-like column -> defer update.
                if (
                    allow_defer_fk
                    and key.endswith("_id")
                    and key in fk_cols_in_plan
                ):
                    transformed[key] = None
                    row_pending.append(
                        PendingUpdate(
                            db_rec=None,
                            column_name=key,
                            placeholder=placeholder,
                            fk_table_name=fk_table_name,
                        )
                    )

            return transformed, id_placeholder, unresolved_pk, row_pending

        with db.same_session() as session:
            # Insert pass: iterate until no progress to handle association rows
            # that depend on previously inserted ids.
            insert_queue = list(to_insert)
            inserted_this_round = True
            while insert_queue and inserted_this_round:
                inserted_this_round = False
                remaining: list[tuple[XlTable[Any], RowDiff]] = []

                for table, row in insert_queue:
                    xl_rec_in = row.xl_row
                    transformed, id_placeholder, unresolved_pk, row_pending = (
                        _transform_xl_rec_for_apply(
                            table,
                            xl_rec_in,
                            allow_defer_fk=True,
                        )
                    )

                    if unresolved_pk:
                        remaining.append((table, row))
                        continue

                    db_rec = table.create_new_db_record(session, transformed)

                    # Bind deferred updates to this newly created record.
                    for pu in row_pending:
                        pu.db_rec = db_rec
                    pending_updates.extend(row_pending)

                    try:
                        # Use a savepoint so one failing row doesn't roll back
                        # prior successful inserts.
                        with session.begin_nested():
                            table.apply_xl_to_db(session, db_rec, transformed)
                            session.add(db_rec)
                            session.flush()
                    except IntegrityError:
                        remaining.append((table, row))
                        continue

                    if id_placeholder is not None:
                        new_id = db_rec.id
                        assert new_id is not None
                        # Store placeholder scoped to this table.
                        placeholder_key = (table.xl_name, id_placeholder)
                        placeholder_to_id[placeholder_key] = new_id

                    result.inserted += 1
                    inserted_this_round = True

                insert_queue = remaining

            if insert_queue:
                # Still unresolved: most likely due to unresolvable placeholders
                # in composite PK association tables.
                unresolved = [(t.xl_name, r.pk) for (t, r) in insert_queue]
                raise ValueError(
                    "Unresolved placeholders prevent inserting %d rows: %r"
                    % (len(unresolved), unresolved)
                )

            # Update pass.
            for table, row in to_update:
                assert row.db_rec is not None
                transformed, _, _, row_pending = _transform_xl_rec_for_apply(
                    table, row.xl_row, allow_defer_fk=True
                )
                for pu in row_pending:
                    pu.db_rec = row.db_rec
                pending_updates.extend(row_pending)
                table.apply_xl_to_db(session, row.db_rec, transformed)
                result.updated += 1

            # Resolve deferred foreign keys now that all inserts ran.
            applied_deferred = 0
            for pu in pending_updates:
                if pu.db_rec is None:
                    continue
                # Look up placeholder scoped to the target table.
                placeholder_key = (pu.fk_table_name, pu.placeholder)
                mapped = placeholder_to_id.get(placeholder_key)
                if mapped is None:
                    continue
                setattr(pu.db_rec, pu.column_name, mapped)
                applied_deferred += 1

            if applied_deferred:
                result.deferred = applied_deferred
                session.flush()

            session.commit()

        # Store placeholder mapping in result for potential Excel file updates.
        result.placeholder_to_id = placeholder_to_id

        return result

    def has_table(self, name: str) -> bool:
        """Returns True if the schema has a table with the given name.

        Args:
            name: Structured table name (`XlTable.xl_name`).
        """
        return any(table.xl_name == name for table in self.tables)

    def get_table(self, name: str) -> "XlTable | None":
        """Returns the table with the given name.

        Args:
            name: Structured table name (`XlTable.xl_name`).

        Returns:
            The matching table instance, or `None` if it does not exist.
        """
        for table in self.tables:
            if table.xl_name == name:
                return table
        return None

    def before_export(self, wb: "Workbook", session: "Session"):
        """Hook called once before exporting tables.

        Args:
            wb: Workbook that will be written to disk.
            session: SQLAlchemy session used for export queries.
        """

    def after_export(self, wb: "Workbook", session: "Session"):
        """Hook called once after exporting tables.

        Args:
            wb: Workbook that will be written to disk.
            session: SQLAlchemy session used for export queries.
        """

    def update_excel_with_allocated_ids(
        self,
        path: str,
        plan: ImportPlan,
        placeholder_to_id: dict[tuple[str, str], int],
        *,
        is_db_pk: Callable[[Any], bool] | None = None,
        wb: Any | None = None,
    ) -> None:
        """Update an Excel file with allocated database IDs.

        This method updates the Excel file in-place, replacing temporary
        placeholder IDs (e.g., "x1", "x2") with the actual integer IDs allocated
        by the database during import.

        Args:
            path: Path to the Excel file to update (used if `wb` is None).
            plan: Import plan that was applied.
            placeholder_to_id: Mapping from (table_name, placeholder_string) to
                allocated integer ID.
            is_db_pk: Predicate used to determine DB IDs. Defaults to
                `_default_is_db_pk`.
            wb: Optional workbook to update. If provided, `path` is only used
                for saving. If None, the workbook is loaded from `path`.
        """
        from openpyxl.utils.cell import range_boundaries  # type: ignore[import]

        if not placeholder_to_id:
            return

        is_db_pk = is_db_pk or _default_is_db_pk
        should_close = False
        if wb is None:
            wb = load_workbook(filename=path, read_only=False, data_only=True)
            should_close = True

        try:
            for table_plan in plan.tables:
                table = table_plan.table
                ws_name = table.sheet_name[0:31]
                if ws_name not in wb.sheetnames:
                    continue
                ws = wb[ws_name]

                ws_table = ws.tables.get(table.xl_name)
                if ws_table is None:
                    continue

                if not ws_table.ref:
                    continue

                # Get table boundaries.
                min_col, min_row, max_col, max_row = range_boundaries(
                    ws_table.ref
                )
                if (
                    min_col is None
                    or min_row is None
                    or max_col is None
                    or max_row is None
                ):
                    continue

                # Exclude Excel "Totals" rows when present.
                totals_row_count = int(
                    getattr(ws_table, "totalsRowCount", 0) or 0
                )
                if totals_row_count:
                    max_row = max(min_row, max_row - totals_row_count)

                # Build column name to index mapping.
                header_row = next(
                    ws.iter_rows(
                        min_row=min_row,
                        max_row=min_row,
                        min_col=min_col,
                        max_col=max_col,
                    )
                )
                col_name_to_idx: dict[str, int] = {}
                for cell in header_row:
                    col_name = str(cell.value) if cell.value else ""
                    if col_name:
                        try:
                            ws_col_idx = table._get_ws_col_idx(cell)
                            col_idx = ws_col_idx - min_col
                            col_name_to_idx[col_name] = col_idx
                        except (TypeError, AttributeError):
                            continue

                # Build column objects lookup.
                col_by_name: dict[str, Any] = {}
                for c in table.get_included_columns():
                    col_by_name[c.xl_name] = c

                # Track which rows had placeholders and need updates.
                rows_to_update: dict[int, dict[str, int]] = {}
                # Map: data row index -> {column_name: new_id}

                # Iterate through all data rows and check for placeholders.
                data_row_idx = 0
                for row in ws.iter_rows(
                    min_row=min_row + 1,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                ):
                    row_updates: dict[str, int] = {}
                    for col_name, col_idx in col_name_to_idx.items():
                        if col_idx >= len(row):
                            continue
                        cell = row[col_idx]
                        value = cell.value

                        # Skip if not a string (placeholders are strings).
                        if not isinstance(value, str):
                            continue

                        placeholder = value.strip()
                        if not placeholder:
                            continue

                        col = col_by_name.get(col_name)
                        if col is None:
                            continue

                        # Check if this is a primary key column with a
                        # placeholder.
                        if col.primary and col_name == "id":
                            placeholder_key = (table.xl_name, placeholder)
                            new_id = placeholder_to_id.get(placeholder_key)
                            if new_id is not None:
                                row_updates[col_name] = new_id
                                continue

                        # Check if this is a foreign key column with a
                        # placeholder.
                        fk_table_name = getattr(col, "fk_table", None)
                        if fk_table_name:
                            placeholder_key = (fk_table_name, placeholder)
                            new_id = placeholder_to_id.get(placeholder_key)
                            if new_id is not None:
                                row_updates[col_name] = new_id

                    if row_updates:
                        rows_to_update[data_row_idx] = row_updates
                    data_row_idx += 1

                # Apply updates to cells.
                for data_row_idx, updates in rows_to_update.items():
                    excel_row = min_row + 1 + data_row_idx
                    for col_name, new_id in updates.items():
                        col_idx_opt = col_name_to_idx.get(col_name)
                        if col_idx_opt is None:
                            continue
                        col_idx = col_idx_opt
                        cell = ws.cell(
                            row=excel_row,
                            column=min_col + col_idx + 1,
                        )
                        # Skip merged cells (they are read-only).
                        # Check by class name to avoid import dependency.
                        if (
                            hasattr(cell, "__class__")
                            and cell.__class__.__name__ == "MergedCell"
                        ):
                            continue
                        # Type checker may complain, but we've checked above.
                        cell.value = new_id  # type: ignore[assignment]

            # Rotate backups before saving.
            from exdrf_util.rotate_backups import (  # type: ignore[import]
                rotate_backups,
            )

            rotate_backups(path)

            wb.save(path)
        finally:
            if should_close:
                wb.close()
