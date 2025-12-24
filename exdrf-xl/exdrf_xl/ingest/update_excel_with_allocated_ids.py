"""Standalone function to update Excel file with allocated database IDs."""

from __future__ import annotations

import logging
import os
from typing import TYPE_CHECKING, Any, Callable

from attrs import define, field
from exdrf_util.rotate_backups import rotate_backups  # type: ignore[import]
from openpyxl.utils.cell import range_boundaries  # type: ignore[import]

from exdrf_xl.ingest.import_plan import ImportPlan
from exdrf_xl.ingest.tools import default_is_db_pk

if TYPE_CHECKING:
    from exdrf_xl.column import XlColumn

logger = logging.getLogger(__name__)


@define
class ExcelIdUpdater:
    """Updates an Excel file with allocated database IDs.

    Replaces temporary placeholder IDs (e.g., "x1", "x2") with the actual
    integer IDs allocated by the database during import.

    Attributes:
        wb: Workbook to update (must be provided).
        plan: Import plan that was applied.
        placeholder_to_id: Mapping from (table_name, placeholder_string) to
            allocated integer ID.
        path: Path to save the workbook to. Required if saving.
        is_db_pk: Predicate used to determine DB IDs.
    """

    wb: Any
    plan: ImportPlan
    placeholder_to_id: dict[tuple[str, str], int]
    path: str | None = None
    is_db_pk: Callable[[Any], bool] = field(
        default=default_is_db_pk,
        converter=lambda x: default_is_db_pk if x is None else x,
    )

    def __call__(self) -> None:
        """Update the Excel file with allocated IDs."""
        if not self.placeholder_to_id:
            return

        self._process_all_tables()
        if self.path is not None:
            self._save_workbook()

    def _process_all_tables(self) -> None:
        """Process all tables in the import plan."""
        assert self.wb is not None
        for table_plan in self.plan.tables:
            self._process_table(table_plan.table)

    def _process_table(self, table: Any) -> None:
        """Process a single table, updating placeholders with allocated IDs.

        Args:
            table: The table to process.
        """
        assert self.wb is not None
        ws = self._get_worksheet(table)
        if ws is None:
            return

        ws_table = ws.tables.get(table.xl_name)
        if ws_table is None:
            return

        boundaries = self._get_table_boundaries(ws_table)
        if boundaries is None:
            return

        min_col, min_row, max_col, max_row = boundaries
        col_name_to_idx = self._build_column_mapping(
            ws, table, min_col, min_row, max_col
        )
        if not col_name_to_idx:
            return

        col_by_name = self._build_column_lookup(table)
        rows_to_update = self._find_placeholder_updates(
            ws,
            table,
            col_name_to_idx,
            col_by_name,
            min_row,
            max_row,
            min_col,
            max_col,
        )

        if rows_to_update:
            self._apply_updates(
                ws, rows_to_update, col_name_to_idx, min_row, min_col
            )

    def _get_worksheet(self, table: Any) -> Any | None:
        """Get the worksheet for a table, if it exists.

        Args:
            table: The table to get worksheet for.

        Returns:
            Worksheet if found, None otherwise.
        """
        assert self.wb is not None
        ws_name = table.sheet_name[0:31]
        if ws_name not in self.wb.sheetnames:
            return None
        return self.wb[ws_name]

    def _get_table_boundaries(
        self, ws_table: Any
    ) -> tuple[int, int, int, int] | None:
        """Get table boundaries from Excel table reference.

        Args:
            ws_table: Excel table object.

        Returns:
            Tuple of (min_col, min_row, max_col, max_row) or None if invalid.
        """
        if not ws_table.ref:
            return None

        min_col, min_row, max_col, max_row = range_boundaries(ws_table.ref)
        if (
            min_col is None
            or min_row is None
            or max_col is None
            or max_row is None
        ):
            return None

        # Exclude Excel "Totals" rows when present.
        totals_row_count = int(getattr(ws_table, "totalsRowCount", 0) or 0)
        if totals_row_count:
            max_row = max(min_row, max_row - totals_row_count)

        return (min_col, min_row, max_col, max_row)

    def _build_column_mapping(
        self,
        ws: Any,
        table: Any,
        min_col: int,
        min_row: int,
        max_col: int,
    ) -> dict[str, int]:
        """Build mapping from column name to column index.

        Args:
            ws: Worksheet.
            table: Table object.
            min_col: Minimum column index.
            min_row: Header row index.
            max_col: Maximum column index.

        Returns:
            Map from column name to relative column index.
        """
        header_row = next(
            ws.iter_rows(
                min_row=min_row,
                max_row=min_row,
                min_col=min_col,
                max_col=max_col,
            )
        )

        col_name_to_idx: dict[str, int] = {}
        for cell in header_row:
            col_name = str(cell.value) if cell.value else ""
            if col_name:
                try:
                    ws_col_idx = table._get_ws_col_idx(cell)
                    col_idx = ws_col_idx - min_col
                    col_name_to_idx[col_name] = col_idx
                except (TypeError, AttributeError):
                    continue

        return col_name_to_idx

    def _build_column_lookup(self, table: Any) -> dict[str, Any]:
        """Build lookup map from column name to column object.

        Args:
            table: Table to build lookup for.

        Returns:
            Map from column name to column object.
        """
        col_by_name: dict[str, Any] = {}
        for c in table.get_included_columns():
            col_by_name[c.xl_name] = c
        return col_by_name

    def _find_placeholder_updates(
        self,
        ws: Any,
        table: Any,
        col_name_to_idx: dict[str, int],
        col_by_name: dict[str, Any],
        min_row: int,
        max_row: int,
        min_col: int,
        max_col: int,
    ) -> dict[int, dict[str, int]]:
        """Find all placeholders that need to be updated with allocated IDs.

        Args:
            ws: Worksheet.
            table: Table object.
            col_name_to_idx: Map from column name to column index.
            col_by_name: Map from column name to column object.
            min_row: Minimum row index.
            max_row: Maximum row index.
            min_col: Minimum column index.
            max_col: Maximum column index.

        Returns:
            Map from data row index to dict of {column_name: new_id}.
        """
        rows_to_update: dict[int, dict[str, int]] = {}
        data_row_idx = 0

        for row in ws.iter_rows(
            min_row=min_row + 1,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            row_updates: dict[str, int] = {}
            for col_name, col_idx in col_name_to_idx.items():
                if col_idx >= len(row):
                    continue

                cell = row[col_idx]
                value = cell.value

                # Skip if not a string (placeholders are strings).
                if not isinstance(value, str):
                    continue

                placeholder = value.strip()
                if not placeholder:
                    continue

                col = col_by_name.get(col_name)
                if col is None:
                    continue

                new_id = self._resolve_placeholder(
                    table, col, col_name, placeholder
                )
                if new_id is not None:
                    row_updates[col_name] = new_id

            if row_updates:
                rows_to_update[data_row_idx] = row_updates
            data_row_idx += 1

        return rows_to_update

    def _resolve_placeholder(
        self, table: Any, col: "XlColumn", col_name: str, placeholder: str
    ) -> int | None:
        """Resolve a placeholder to an allocated ID.

        Args:
            table: Table object.
            col: Column object.
            col_name: Column name.
            placeholder: Placeholder string.

        Returns:
            Allocated ID if found, None otherwise.
        """
        # Check if this is a primary key column with a placeholder.
        if col.primary and col_name == "id":
            placeholder_key = (table.xl_name, placeholder)
            return self.placeholder_to_id.get(placeholder_key)

        # Check if this is a foreign key column with a placeholder.
        fk_table_name = col.fk_table
        if fk_table_name:
            placeholder_key = (fk_table_name, placeholder)
            return self.placeholder_to_id.get(placeholder_key)

        return None

    def _apply_updates(
        self,
        ws: Any,
        rows_to_update: dict[int, dict[str, int]],
        col_name_to_idx: dict[str, int],
        min_row: int,
        min_col: int,
    ) -> None:
        """Apply updates to Excel cells.

        Args:
            ws: Worksheet.
            rows_to_update: Map from data row index to {column_name: new_id}.
            col_name_to_idx: Map from column name to column index.
            min_row: Minimum row index (header row).
            min_col: Minimum column index.
        """
        for data_row_idx, updates in rows_to_update.items():
            excel_row = min_row + 1 + data_row_idx
            for col_name, new_id in updates.items():
                col_idx_opt = col_name_to_idx.get(col_name)
                if col_idx_opt is None:
                    continue

                cell = ws.cell(
                    row=excel_row,
                    column=min_col + col_idx_opt,
                )

                # Skip merged cells (they are read-only).
                # Check by class name to avoid import dependency.
                if (
                    hasattr(cell, "__class__")
                    and cell.__class__.__name__ == "MergedCell"
                ):
                    continue

                # Type checker may complain, but we've checked above.
                cell.value = new_id  # type: ignore[assignment]

    def _save_workbook(self) -> None:
        """Save the workbook and rotate backups."""
        assert self.path is not None

        crt_base, ext = os.path.splitext(self.path)
        counter = -1

        while counter < 100:
            counter += 1
            if counter == 0:
                crt_path = self.path
            elif counter == 50:
                base_name = os.path.splitext(os.path.basename(self.path))[0]

                import tempfile

                crt_base = os.path.join(tempfile.gettempdir(), base_name)
                crt_path = f"{crt_base}-{ext}{ext}"

                logger.warning(
                    "Unable to save the updated file in same location "
                    "as the original file (`%s`). Will attempt to use the "
                    "temporary directory `%s`",
                    self.path,
                    crt_path,
                )
            else:
                crt_path = f"{crt_base}-{ext}{ext}"

            rotate_backups(crt_path)
            try:
                self.wb.save(crt_path)
                if counter > 0:
                    logger.info(
                        "Updated file was saved as `%s` because original "
                        "path was not writable",
                        crt_path,
                    )
                break
            except PermissionError:
                pass


def update_excel_with_allocated_ids(
    wb: Any,
    plan: ImportPlan,
    placeholder_to_id: dict[tuple[str, str], int],
    *,
    path: str | None = None,
    is_db_pk: Callable[[Any], bool] | None = None,
) -> None:
    """Update an Excel file with allocated database IDs.

    This function updates the workbook in-place, replacing temporary
    placeholder IDs (e.g., "x1", "x2") with the actual integer IDs allocated
    by the database during import.

    Args:
        wb: Workbook to update (must be provided, already loaded).
        plan: Import plan that was applied.
        placeholder_to_id: Mapping from (table_name, placeholder_string) to
            allocated integer ID.
        path: Optional path to save the workbook to. If provided, the workbook
            will be saved and backups rotated. If None, updates are made in
            memory only.
        is_db_pk: Predicate used to determine DB IDs. Defaults to
            `default_is_db_pk`.
    """
    updater = ExcelIdUpdater(
        wb, plan, placeholder_to_id, path=path, is_db_pk=is_db_pk
    )
    updater()
