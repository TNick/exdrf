"""Standalone function to build an import plan from an Excel workbook."""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING, Any, Callable

from attrs import define, field
from openpyxl import load_workbook  # type: ignore[import]
from sqlalchemy.exc import DataError

from exdrf_xl.ingest.cell_diff import CellDiff
from exdrf_xl.ingest.import_plan import ImportPlan
from exdrf_xl.ingest.row_diff import RowDiff, XlRecord
from exdrf_xl.ingest.table_diff import TableDiff
from exdrf_xl.ingest.tools import (
    default_is_db_pk,
    is_datetime_type,
    iter_chunks,
    normalize_unknown_datetime,
)

if TYPE_CHECKING:
    from exdrf_xl.schema import XlSchema
    from exdrf_xl.table import XlTable

logger = logging.getLogger(__name__)


@define
class ImportPlanBuilder:
    """Builds an import plan by comparing Excel data with database records.

    This class inspects each configured `XlTable` present in the workbook and
    classifies each row as:
    - new row (primary key looks like a placeholder), or
    - existing row (primary key looks like a DB id and can be found), or
    - modified existing row (existing row with changed values)

    Attributes:
        schema: The XlSchema instance containing table definitions.
        db: Database connection used to create a session for comparisons.
        path: Input `.xlsx` file path (used if `wb` is None).
        is_db_pk: Predicate used to determine whether a primary key cell comes
            from the database.
        batch_size: Batch size for database queries.
        wb: Optional workbook to use. If provided, `path` is only used for
            reference. If None, the workbook is loaded from `path`.
        should_close: Whether to close the workbook when done.
    """

    schema: "XlSchema"
    db: Any
    path: str
    is_db_pk: Callable[[Any], bool] = field(
        default=default_is_db_pk,
        converter=lambda x: default_is_db_pk if x is None else x,
    )
    batch_size: int = 100
    wb: Any | None = None
    should_close: bool = field(default=False, init=False)

    def __call__(self) -> ImportPlan:
        """Build and return the import plan.

        Returns:
            An `ImportPlan` describing all detected new/modified rows.
        """
        self._ensure_workbook()
        try:
            with self.db.same_session() as session:
                table_plans: list[TableDiff] = []
                for table in self.schema.tables:
                    plan = self._process_table(table, session)
                    if plan is not None:
                        table_plans.append(plan)
                return ImportPlan(
                    source_path=self.path, tables=tuple(table_plans)
                )
        finally:
            self._close_workbook()

    def _ensure_workbook(self) -> None:
        """Load workbook if not provided."""
        if self.wb is None:
            self.wb = load_workbook(
                filename=self.path, read_only=False, data_only=True
            )
            self.should_close = True

    def _close_workbook(self) -> None:
        """Close workbook if we opened it."""
        if self.should_close and self.wb is not None:
            self.wb.close()

    def _process_table(
        self, table: "XlTable[Any]", session: Any
    ) -> TableDiff | None:
        """Process a single table and return its diff, if any changes found.

        Args:
            table: The table to process.
            session: Database session for lookups.

        Returns:
            TableDiff if changes found, None otherwise.
        """
        ws = self._get_worksheet(table)
        if ws is None:
            logger.error("Unable to locate table `%s` worksheet", table.xl_name)
            return None

        ws_table = ws.tables.get(table.xl_name)
        if ws_table is None:
            logger.error("No such Excel table `%s`", table.xl_name)
            return None

        xl_rows = list(table.iter_excel_table(ws, ws_table))
        if not xl_rows:
            logger.debug("Table `%s` has no rows", table.xl_name)
            return None

        processor = TableRowProcessor(
            table, xl_rows, session, self.is_db_pk, self.batch_size
        )
        new_rows, modified_rows, existing_rows = processor.process()

        if new_rows or modified_rows:
            return TableDiff(
                table=table,
                new_rows=tuple(new_rows),
                modified_rows=tuple(modified_rows),
                existing_rows=existing_rows,
                total_rows=len(xl_rows),
            )
        return None

    def _get_worksheet(self, table: "XlTable[Any]") -> Any | None:
        """Get the worksheet for a table, if it exists.

        Args:
            table: The table to get worksheet for.

        Returns:
            Worksheet if found, None otherwise.
        """
        assert self.wb is not None
        ws_name = table.sheet_name[0:31]
        if ws_name not in self.wb.sheetnames:
            return None
        return self.wb[ws_name]


class TableRowProcessor:
    """Processes rows for a single table to identify new/modified rows."""

    def __init__(
        self,
        table: "XlTable[Any]",
        xl_rows: list[XlRecord],
        session: Any,
        is_db_pk: Callable[[Any], bool],
        batch_size: int,
    ):
        """Initialize the row processor.

        Args:
            table: The table being processed.
            xl_rows: All Excel rows for this table.
            session: Database session for lookups.
            is_db_pk: Predicate to determine if a value is a DB primary key.
            batch_size: Batch size for database queries.
        """
        self.table = table
        self.xl_rows = xl_rows
        self.session = session
        self.is_db_pk = is_db_pk
        self.batch_size = batch_size

        self.pk_cols = [c.xl_name for c in table.columns if c.primary]
        self.pk_cols_eff = (
            table.get_pk_column_names()
            if hasattr(table, "get_pk_column_names")
            else self.pk_cols
        )

        self.new_rows: list[RowDiff] = []
        self.modified_rows: list[RowDiff] = []
        self.existing_rows = 0

    def process(
        self,
    ) -> tuple[list[RowDiff], list[RowDiff], int]:
        """Process all rows and return categorized results.

        Returns:
            Tuple of (new_rows, modified_rows, existing_rows_count).
        """
        db_model = self._get_db_model()
        if db_model is not None and self.pk_cols_eff:
            self._process_batched(db_model)
        else:
            self._process_legacy()
        return self.new_rows, self.modified_rows, self.existing_rows

    def _get_db_model(self) -> Any | None:
        """Get the database model class for this table, if available.

        Returns:
            Model class or None if not available.
        """
        try:
            return self.table.get_db_model_class()
        except Exception:
            return None

    def _process_batched(self, db_model: Any) -> None:
        """Process rows using batched database lookups (preferred path).

        Args:
            db_model: The SQLAlchemy model class for this table.
        """
        pk_key_to_xl, keyed_conditions = self._build_lookup_maps()
        found_keys = self._execute_batched_lookups(
            db_model, pk_key_to_xl, keyed_conditions
        )
        self._handle_unfound_keys(keyed_conditions, found_keys, pk_key_to_xl)

    def _build_lookup_maps(
        self,
    ) -> tuple[
        dict[tuple[Any, ...], XlRecord],
        list[tuple[tuple[Any, ...], tuple[Any, ...]]],
    ]:
        """Build maps for batched lookup.

        Returns:
            Tuple of (pk_key_to_xl_record, keyed_conditions).
        """
        pk_key_to_xl: dict[tuple[Any, ...], XlRecord] = {}
        keyed_conditions: list[tuple[tuple[Any, ...], tuple[Any, ...]]] = []

        for xl_rec in self.xl_rows:
            pk_key = tuple(xl_rec.get(k, None) for k in self.pk_cols_eff)
            conds = self._get_pk_conditions(xl_rec)

            if conds is None:
                self._add_new_row(xl_rec)
                continue

            pk_key_to_xl[pk_key] = xl_rec
            keyed_conditions.append((pk_key, conds))

        return pk_key_to_xl, keyed_conditions

    def _get_pk_conditions(self, xl_rec: XlRecord) -> tuple[Any, ...] | None:
        """Get primary key conditions for an Excel record.

        Args:
            xl_rec: Excel record to get conditions for.

        Returns:
            Conditions tuple or None if this looks like a new row.
        """
        try:
            return self.table.pk_conditions(xl_rec, self.is_db_pk)
        except Exception:
            logger.exception(
                "PK conditions failed for table %s", self.table.xl_name
            )
            return None

    def _execute_batched_lookups(
        self,
        db_model: Any,
        pk_key_to_xl: dict[tuple[Any, ...], XlRecord],
        keyed_conditions: list[tuple[tuple[Any, ...], tuple[Any, ...]]],
    ) -> set[tuple[Any, ...]]:
        """Execute batched database lookups.

        Args:
            db_model: The SQLAlchemy model class.
            pk_key_to_xl: Map from PK keys to Excel records.
            keyed_conditions: List of (pk_key, conditions) tuples.

        Returns:
            Set of found primary key tuples.
        """
        found_keys: set[tuple[Any, ...]] = set()
        for chunk in iter_chunks(keyed_conditions, self.batch_size):
            try:
                found_keys.update(
                    self._process_batch_chunk(db_model, chunk, pk_key_to_xl)
                )
            except DataError:
                # Postgres aborts on DataError; rollback and fallback.
                self.session.rollback()
                found_keys.update(
                    self._process_batch_chunk_fallback(chunk, pk_key_to_xl)
                )
        return found_keys

    def _process_batch_chunk(
        self,
        db_model: Any,
        chunk: list[tuple[tuple[Any, ...], tuple[Any, ...]]],
        pk_key_to_xl: dict[tuple[Any, ...], XlRecord],
    ) -> set[tuple[Any, ...]]:
        """Process a single batch chunk via SQL query.

        Args:
            db_model: The SQLAlchemy model class.
            chunk: Chunk of (pk_key, conditions) tuples.
            pk_key_to_xl: Map from PK keys to Excel records.

        Returns:
            Set of found primary key tuples.
        """
        from sqlalchemy import and_, or_, select

        where_clause = or_(*[and_(*conds) for _, conds in chunk])
        stmt = select(db_model).where(where_clause)

        found_keys: set[tuple[Any, ...]] = set()
        for db_rec in self.session.scalars(stmt):
            db_key = tuple(getattr(db_rec, k) for k in self.pk_cols_eff)
            xl_rec_opt = pk_key_to_xl.get(db_key)
            if xl_rec_opt is not None:
                found_keys.add(db_key)
                self._add_existing_row(db_rec, xl_rec_opt)
        return found_keys

    def _process_batch_chunk_fallback(
        self,
        chunk: list[tuple[tuple[Any, ...], tuple[Any, ...]]],
        pk_key_to_xl: dict[tuple[Any, ...], XlRecord],
    ) -> set[tuple[Any, ...]]:
        """Fallback: process chunk via per-row lookups.

        Args:
            chunk: Chunk of (pk_key, conditions) tuples.
            pk_key_to_xl: Map from PK keys to Excel records.

        Returns:
            Set of found primary key tuples.
        """
        found_keys: set[tuple[Any, ...]] = set()
        for pk_key, _conds in chunk:
            xl_rec_opt = pk_key_to_xl.get(pk_key)
            if xl_rec_opt is None:
                continue

            db_rec = None
            try:
                db_rec = self.table.find_db_rec(
                    self.session, xl_rec_opt, self.is_db_pk
                )
            except DataError:
                self.session.rollback()
                db_rec = None

            if db_rec is None:
                self._add_new_row(xl_rec_opt)
            else:
                found_keys.add(pk_key)
                self._add_existing_row(db_rec, xl_rec_opt)
        return found_keys

    def _handle_unfound_keys(
        self,
        keyed_conditions: list[tuple[tuple[Any, ...], tuple[Any, ...]]],
        found_keys: set[tuple[Any, ...]],
        pk_key_to_xl: dict[tuple[Any, ...], XlRecord],
    ) -> None:
        """Handle primary keys that looked like DB IDs but weren't found.

        Args:
            keyed_conditions: All (pk_key, conditions) tuples.
            found_keys: Set of keys that were found in the database.
            pk_key_to_xl: Map from PK keys to Excel records.
        """
        for pk_key, _conds in keyed_conditions:
            if pk_key in found_keys:
                continue
            xl_rec_opt = pk_key_to_xl.get(pk_key)
            if xl_rec_opt is not None:
                self._add_new_row(xl_rec_opt)

    def _process_legacy(self) -> None:
        """Process rows using legacy per-row lookup (fallback path)."""
        for xl_rec in self.xl_rows:
            pk_is_db = self.pk_cols_eff and all(
                self.is_db_pk(xl_rec.get(k)) for k in self.pk_cols_eff
            )

            db_rec = None
            if pk_is_db:
                try:
                    db_rec = self.table.find_db_rec(
                        self.session, xl_rec, self.is_db_pk
                    )
                except DataError:
                    self.session.rollback()
                    db_rec = None

            if db_rec is None:
                self._add_new_row(xl_rec)
            else:
                self._add_existing_row(db_rec, xl_rec)

    def _add_new_row(self, xl_rec: XlRecord) -> None:
        """Add a new row diff for a row that doesn't exist in the database.

        Args:
            xl_rec: Excel record for the new row.
        """
        pk = {k: xl_rec.get(k) for k in self.pk_cols_eff if k in xl_rec}
        diffs = tuple(
            CellDiff(
                column=c.xl_name,
                old_value=None,
                new_value=xl_rec.get(c.xl_name, None),
            )
            for c in self.table.columns
            if not bool(getattr(c, "read_only", False))
            if c.xl_name in xl_rec
        )
        self.new_rows.append(
            RowDiff(
                table_name=self.table.xl_name,
                is_new=True,
                pk=pk,
                xl_row=xl_rec,
                db_rec=None,
                diffs=diffs,
            )
        )

    def _add_existing_row(self, db_rec: Any, xl_rec: XlRecord) -> None:
        """Add an existing row diff, checking for modifications.

        Args:
            db_rec: Database record.
            xl_rec: Excel record to compare against.
        """
        self.existing_rows += 1
        pk = {k: xl_rec.get(k) for k in self.pk_cols_eff if k in xl_rec}
        cell_diffs = self._compute_cell_diffs(db_rec, xl_rec)

        if cell_diffs:
            self.modified_rows.append(
                RowDiff(
                    table_name=self.table.xl_name,
                    is_new=False,
                    pk=pk,
                    xl_row=xl_rec,
                    db_rec=db_rec,
                    diffs=tuple(cell_diffs),
                )
            )

    def _compute_cell_diffs(
        self, db_rec: Any, xl_rec: XlRecord
    ) -> list[CellDiff]:
        """Compute differences between database record and Excel record.

        Args:
            db_rec: Database record.
            xl_rec: Excel record.

        Returns:
            List of cell differences.
        """
        cell_diffs: list[CellDiff] = []
        for c in self.table.columns:
            if bool(getattr(c, "read_only", False)):
                continue
            if c.xl_name not in xl_rec:
                continue

            new_val = xl_rec.get(c.xl_name, None)
            old_val = c.value_from_record(db_rec)

            # Normalize unknown datetime sentinel so we don't report diffs
            # due to Excel range limitations.
            if is_datetime_type(getattr(c, "type_name", None)):
                new_val = normalize_unknown_datetime(new_val)
                old_val = normalize_unknown_datetime(old_val)

            # Normalize string whitespace.
            if isinstance(new_val, str):
                new_val = new_val.strip()
            if isinstance(old_val, str):
                old_val = old_val.strip()

            if old_val != new_val:
                cell_diffs.append(
                    CellDiff(
                        column=c.xl_name,
                        old_value=old_val,
                        new_value=new_val,
                    )
                )
        return cell_diffs


def plan_import_from_file(
    schema: "XlSchema",
    db: Any,
    path: str,
    *,
    is_db_pk: Callable[[Any], bool] | None = None,
    batch_size: int = 100,
    wb: Any | None = None,
) -> ImportPlan:
    """Build an import plan from an Excel workbook.

    This inspects each configured `XlTable` present in the workbook and
    classifies each row as:
    - new row (primary key looks like a placeholder), or
    - existing row (primary key looks like a DB id and can be found), or
    - modified existing row (existing row with changed values)

    Args:
        schema: The XlSchema instance containing table definitions.
        db: Database connection used to create a session for comparisons.
        path: Input `.xlsx` file path (used if `wb` is None).
        is_db_pk: Optional predicate used to determine whether a primary key
            cell comes from the database. Defaults to `default_is_db_pk`.
        batch_size: Batch size for database queries.
        wb: Optional workbook to use. If provided, `path` is only used for
            reference. If None, the workbook is loaded from `path`.

    Returns:
        An `ImportPlan` describing all detected new/modified rows.
    """
    builder = ImportPlanBuilder(
        schema, db, path, is_db_pk=is_db_pk, batch_size=batch_size, wb=wb
    )
    return builder()
