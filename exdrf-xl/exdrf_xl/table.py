import logging
from copy import copy
from datetime import datetime
from typing import (
    TYPE_CHECKING,
    Any,
    Callable,
    Dict,
    Generator,
    Generic,
    Mapping,
    TypeVar,
    cast,
)

from attrs import define, field
from exdrf.constants import FIELD_TYPE_DT  # type: ignore[import]
from exdrf.field_types.date_time import UNKNOWN_DATETIME
from openpyxl.formatting.rule import Rule  # type: ignore[import]
from openpyxl.styles import (  # type: ignore[import]
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.styles.colors import Color  # type: ignore[import]
from openpyxl.styles.differential import (  # type: ignore[import]
    DifferentialStyle,
)
from openpyxl.utils import get_column_letter  # type: ignore[import]
from openpyxl.utils.cell import (  # type: ignore[import]
    column_index_from_string,
    range_boundaries,
)
from openpyxl.worksheet.table import (  # type: ignore[import]
    Table,
    TableColumn,
    TableStyleInfo,
)
from sqlalchemy import func, select

from exdrf_xl.utils.parse_int import parse_int
from exdrf_xl.utils.rgb import normalize_rgb_color
from exdrf_xl.utils.write_cell import install_custom_lxml_writer

if TYPE_CHECKING:
    from openpyxl.cell import Cell
    from openpyxl.worksheet.worksheet import Worksheet
    from sqlalchemy.orm import Session

    from exdrf_xl.column import XlColumn


T = TypeVar("T")
logger = logging.getLogger(__name__)
install_custom_lxml_writer()


@define
class XlTable(Generic[T]):
    """Describes the mapping between a database table and an
    Excel sheet.

    Attributes:
        schema: The schema that contains the table.
        sheet_name: Worksheet name to create in the workbook.
        xl_name: Structured table name (Excel "Table" displayName).
        columns: Column definitions in display order.

            Notes:
            - Only columns where `XlColumn.is_included()` returns `True` are
              written.
            - Column formatting (widths, alignments, optional font/fill colors)
              is applied to data rows (starting at row 2), not the header row.
        _included_columns_cache: Cached list of included columns (those where
            `XlColumn.is_included()` is `True`). This is an internal cache.
        _included_index_by_name_cache: Cached mapping from column name
            (`XlColumn.xl_name`) to 0-based index in the included columns list.
            This is an internal cache.
        _included_index_by_objid_cache: Cached mapping from column object id
            (`id(column)`) to 0-based index in the included columns list. This
            is an internal cache.
    """

    schema: Any = field(repr=False)
    sheet_name: str
    xl_name: str
    columns: list["XlColumn"] = field(factory=list, repr=False)
    pk_columns: tuple[str, ...] = field(default=None, init=False, repr=False)
    _included_columns_cache: list[Any] | None = field(
        default=None, init=False, repr=False
    )
    _included_index_by_name_cache: dict[str, int] | None = field(
        default=None, init=False, repr=False
    )
    _included_index_by_objid_cache: dict[int, int] | None = field(
        default=None, init=False, repr=False
    )

    def __attrs_post_init__(self):
        """Finalize initialization after attrs has constructed the instance.

        This wires each column definition back to this table and initializes
        the included-column caches.
        """
        for c in self.columns:
            c.table = self
        self._invalidate_column_caches()

    def get_pk_column_names(self) -> tuple[str, ...]:
        """Return the primary key column names in stable order.

        Generated table classes can optionally define a `pk_columns` attribute.
        If not available, we fall back to scanning `self.columns` for columns
        marked as `primary` (in definition order).
        """
        pk_cols = self.pk_columns
        if isinstance(pk_cols, (tuple, list)) and len(pk_cols) > 0:
            return tuple(pk_cols)

        result = tuple(c.xl_name for c in self.columns if c.primary)
        self.pk_columns = result
        return result

    def get_db_model_class(self) -> Any:
        """Return the SQLAlchemy model class for this table.

        Generated classes should override this so schema planning can batch
        lookups without relying on reflection/inspection.
        """
        raise NotImplementedError

    def build_pk_conditions(
        self,
        db_model: Any,
        xl_rec: Dict[str, Any],
        is_db_pk: Callable[[Any], bool],
    ) -> tuple[Any, ...] | None:
        """Build SQLAlchemy WHERE conditions that identify a DB record.

        Args:
            db_model: SQLAlchemy mapped class (e.g. `DbMyTable`).
            xl_rec: Excel row dictionary.
            is_db_pk: Predicate used to decide whether a PK value is from the
                database (True) or is a placeholder for new records (False).

        Returns:
            Tuple of SQLAlchemy binary expressions suitable for
            `select(...).where(*conditions)`, or `None` if the record cannot be
            matched to an existing DB row.
        """
        pk_cols = self.get_pk_column_names()
        if not pk_cols:
            return None

        conditions: list[Any] = []
        for col_name in pk_cols:
            v = xl_rec.get(col_name, None)
            if v is None or (isinstance(v, str) and v.strip() == ""):
                return None

            col_def = self[col_name]
            assert col_def is not None
            type_name = col_def.type_name

            # If this PK component is an integer, it must be a DB id
            # (or parse-able as one). Placeholders like "x1" must not be used
            # in comparisons against integer columns.
            if type_name == "integer":
                parsed = parse_int(v)
                if parsed is None:
                    return None
                if not is_db_pk(parsed):
                    return None
                conditions.append(getattr(db_model, col_name) == parsed)
                continue

            # For non-integer PKs (e.g. string PK components), we accept the
            # value as-is (after trimming strings).
            if isinstance(v, str):
                v = v.strip()
                if v == "":
                    return None

            conditions.append(getattr(db_model, col_name) == v)

        return tuple(conditions)

    def pk_conditions(
        self,
        xl_rec: Dict[str, Any],
        is_db_pk: Callable[[Any], bool],
    ) -> tuple[Any, ...] | None:
        """Convenience wrapper to build PK match conditions for this table."""
        db_model = self.get_db_model_class()
        return self.build_pk_conditions(db_model, xl_rec, is_db_pk)

    def _invalidate_column_caches(self) -> None:
        """Invalidate cached included-columns data structures.

        This should be called after mutating `columns` or when any column's
        `is_included()` behavior can change.
        """
        self._included_columns_cache = None
        self._included_index_by_name_cache = None
        self._included_index_by_objid_cache = None

    def _ensure_column_caches(self) -> None:
        """Build cached included-columns data structures if missing.

        The caches store:
        - The list of included columns (in display order).
        - Fast lookup by column name and by column object identity.
        """
        if (
            self._included_columns_cache is not None
            and self._included_index_by_name_cache is not None
            and self._included_index_by_objid_cache is not None
        ):
            return

        included = [c for c in self.columns if c.is_included()]
        index_by_name: dict[str, int] = {}
        index_by_objid: dict[int, int] = {}

        for idx, c in enumerate(included):
            # Keep the first match for a name, consistent with prior behavior.
            index_by_name.setdefault(c.xl_name, idx)
            index_by_objid[id(c)] = idx

        self._included_columns_cache = included
        self._included_index_by_name_cache = index_by_name
        self._included_index_by_objid_cache = index_by_objid

    def get_selector(self) -> Any:
        """Returns the SQLAlchemy selector for the table."""
        raise NotImplementedError("Subclasses must implement this method")

    def get_rows(self, session: "Session") -> Generator[T, None, None]:
        """Returns the rows of the table from a database session.

        Args:
            session: SQLAlchemy session used to execute the selector.
        """
        for row in session.scalars(self.get_selector()):
            yield row

    def get_rows_count(self, session: "Session") -> int:
        """Returns the number of rows in the table.

        Args:
            session: SQLAlchemy session used to execute the selector.
        """
        # Count rows from the selector by wrapping it as a subquery so it
        # becomes a valid FROM clause (and satisfies static type checkers).
        subq = self.get_selector().subquery()
        result = session.scalar(select(func.count()).select_from(subq))

        if result is None:
            logger.error(
                "Failed to retrieve rows count for table %s", self.xl_name
            )
            return 0

        return result

    def write_to_sheet(
        self,
        sheet: "Worksheet",
        session: "Session",
        col_widths: Mapping[str, float | None],
    ):
        """Creates the table in the sheet.

        Args:
            sheet: Target worksheet.
            session: SQLAlchemy session used to load rows for export.
        """
        if len(self.columns) == 0:
            logger.warning("Table %s has no columns", self.xl_name)
            # return

        included = self.get_included_columns()

        # Format header row.
        sheet.row_dimensions[1].height = 30
        header_align = Alignment(
            wrap_text=False,
            horizontal="center",
            vertical="top",
        )
        for c, column in enumerate(included):
            cell = sheet.cell(row=1, column=c + 1, value=column.xl_name)
            cell.alignment = header_align

        # Generate data rows.
        row_count = 0
        for row_idx, record in enumerate(self.get_rows(session)):
            row_count += 1
            for column in included:
                # Write data rows starting at row 2 (row 1 is the header).
                column.write_to_sheet(sheet, row_idx, record)

        table_obj = self.create_excel_table(row_count)
        assert table_obj
        sheet.add_table(table_obj)

        self.apply_column_widths(sheet, widths=col_widths)
        self.apply_alignments(sheet)
        self.apply_cell_styles(sheet, row_count)
        self.apply_duplicate_id_conditional_formatting(sheet, row_count)

        for column in included:
            column.post_table_created(sheet, table_obj, row_count)

    def create_excel_table(self, row_count: int) -> "Table":
        """Create the Excel structured table object.

        Args:
            row_count: Number of data rows (excluding header).

        Returns:
            The created structured table object.
        """

        columns = self.get_included_columns()
        num_cols = len(columns)
        last_col = get_column_letter(num_cols)
        num_rows = row_count + 1  # include header row
        ref = f"A1:{last_col}{num_rows}"

        table_obj = Table(displayName=self.xl_name, ref=ref)
        table_obj._initialise_columns()  # type: ignore

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table_obj.tableStyleInfo = style
        table_obj.tableColumns = [
            TableColumn(id=i + 1, name=h.xl_name) for i, h in enumerate(columns)
        ]
        return table_obj

    def apply_column_widths(
        self, ws: "Worksheet", widths: Mapping[str, float | None]
    ):
        """Apply saved column widths from previous export.

        Args:
            ws: Excel worksheet object.
            widths: Dictionary mapping column names to widths.
        """
        # Apply widths to all columns.
        for col_idx, col_def in enumerate(self.get_included_columns(), start=1):
            col_name = col_def.xl_name
            col_letter = get_column_letter(col_idx)
            col = ws.column_dimensions[col_letter]

            # Hide columns that are marked as hidden.
            col.hidden = col_def.hidden

            override_width = widths.get(col_name, None)
            if override_width is None:
                col.width = col_def.col_width
            else:
                col.width = override_width

    def apply_alignments(self, ws: "Worksheet"):
        """Apply column alignment settings to data cells.

        Args:
            ws: Excel worksheet object.
        """
        for col_idx, col_def in enumerate(self.get_included_columns(), start=1):
            align = Alignment(
                wrap_text=col_def.wrap_text,
                horizontal=col_def.h_align,
                vertical=col_def.v_align,
            )
            for row in ws.iter_rows(
                min_col=col_idx, max_col=col_idx, min_row=2
            ):
                row[0].alignment = align

    def apply_cell_styles(self, ws: "Worksheet", row_count: int):
        """Apply per-column cell styles (font color and background fill).

        Args:
            ws: Excel worksheet object.
            row_count: Number of data rows (excluding the header).
        """
        if row_count <= 0:
            return

        max_row = row_count + 1  # include header row at 1; data starts at 2

        for col_idx, col_def in enumerate(self.get_included_columns(), start=1):
            if col_def.font_color is None and col_def.bg_color is None:
                continue

            font_color = None
            if col_def.font_color is not None:
                font_color = Color(rgb=normalize_rgb_color(col_def.font_color))

            fill = None
            if col_def.bg_color is not None:
                rgb = normalize_rgb_color(col_def.bg_color)
                fill = PatternFill(
                    fill_type="solid",
                    start_color=rgb,
                    end_color=rgb,
                )

            for row in ws.iter_rows(
                min_col=col_idx,
                max_col=col_idx,
                min_row=2,
                max_row=max_row,
            ):
                cell = row[0]

                if font_color is not None:
                    # `cell.font` is a StyleProxy. Copy to a real Font and
                    # reassign so we don't mutate the proxy.
                    new_font = cast(Font, copy(cell.font))
                    new_font.color = font_color  # type: ignore
                    cell.font = new_font

                if fill is not None:
                    cell.fill = fill

    def apply_duplicate_id_conditional_formatting(
        self, ws: "Worksheet", row_count: int
    ):
        """Highlight duplicate values in columns named `id`.

        This is a convenience wrapper around
        `apply_duplicate_values_conditional_formatting()` that targets columns
        whose `xl_name` is `"id"` (case-insensitive).

        Args:
            ws: Excel worksheet object.
            row_count: Number of data rows (excluding the header).
        """
        self.apply_duplicate_values_conditional_formatting(
            ws=ws,
            row_count=row_count,
            predicate=lambda c: c.xl_name.strip().lower() == "id",
            fill_color="FFFFC7CE",
            font_color="FFFF0000",
        )

    def apply_duplicate_values_conditional_formatting(
        self,
        ws: "Worksheet",
        row_count: int,
        predicate: Callable[[Any], bool],
        fill_color: str,
        font_color: str,
    ):
        """Add Excel "duplicate values" conditional formatting for columns.

        This adds a native Excel conditional formatting rule of type
        `"duplicateValues"` to each included column that matches `predicate`.
        The rule is applied to the current data range (rows 2..N).

        Args:
            ws: Excel worksheet object.
            row_count: Number of data rows (excluding the header).
            predicate: Function that receives an included `XlColumn` and returns
                True if the rule should be applied to that column.
            fill_color: Background fill color (ARGB hex, e.g. `"FFFFC7CE"`).
            font_color: Font color (ARGB hex, e.g. `"FFFF0000"`).
        """
        if row_count <= 0:
            return

        max_row = row_count + 1  # include header row at 1; data starts at 2

        fill = PatternFill(
            fill_type="solid",
            start_color=fill_color,
            end_color=fill_color,
        )
        font = Font(color=font_color)
        dxf = DifferentialStyle(font=font, fill=fill)

        for col_idx, col_def in enumerate(self.get_included_columns(), start=1):
            if not predicate(col_def):
                continue

            col_letter = get_column_letter(col_idx)

            # We assume here that the table is inserted at A1 and first row
            # is the title row.
            range_ref = f"{col_letter}2:{col_letter}{max_row}"

            rule = Rule(type="duplicateValues")
            rule.dxf = dxf
            ws.conditional_formatting.add(range_ref, rule)

    def __getitem__(self, key: int | str) -> Any:
        """Get an included column by index or by name.

        Args:
            key: Either a 0-based index into the included columns list, or the
                column name (`XlColumn.xl_name`).

        Returns:
            The matching included column.

        Raises:
            IndexError: If an integer index is out of range.
            KeyError: If a string name does not match any included column.
            TypeError: If `key` is not an `int` or `str`.
        """
        self._ensure_column_caches()

        if isinstance(key, int):
            assert self._included_columns_cache is not None
            return self._included_columns_cache[key]

        if isinstance(key, str):
            assert self._included_index_by_name_cache is not None
            idx = self._included_index_by_name_cache.get(key, None)
            if idx is None:
                raise KeyError(key)
            assert self._included_columns_cache is not None
            return self._included_columns_cache[idx]

        raise TypeError("Invalid key type %r; expected int or str" % type(key))

    def __len__(self) -> int:
        return len(self.columns)

    def get_included_columns(self) -> list[Any]:
        """Returns the columns that are included in the table."""
        self._ensure_column_caches()
        assert self._included_columns_cache is not None
        return self._included_columns_cache

    def get_column_index(self, column: "XlColumn | str") -> int:
        """Returns the index of the column in the table.

        Args:
            column: Column reference, either by `XlColumn` instance (identity)
                or by column name (`XlColumn.xl_name`).

        Returns:
            0-based index in the *included* columns list, or -1 if not found.
        """
        self._ensure_column_caches()
        assert self._included_index_by_name_cache is not None
        assert self._included_index_by_objid_cache is not None

        if isinstance(column, str):
            return self._included_index_by_name_cache.get(column, -1)
        else:
            return self._included_index_by_objid_cache.get(id(column), -1)

    @staticmethod
    def _get_ws_col_idx(cell: "Cell") -> int:
        """Return the 1-based worksheet column index for an openpyxl cell.

        Some openpyxl cell-like classes (e.g. `MergedCell`) do not expose
        `col_idx` in their type information, even though a column index can
        always be derived via `column`.
        """
        column = cell.column
        if isinstance(column, int):
            return column
        if isinstance(column, str):
            return column_index_from_string(column)

        raise TypeError(
            "Unsupported cell column type %r for cell %r" % (type(column), cell)
        )

    def iter_excel_table(self, ws: "Worksheet", table: "Table"):
        """Iterate rows from an Excel structured table and yield row dicts.

        Args:
            ws: Worksheet that contains the structured table.
            table: The openpyxl structured table object to iterate.

        Yields:
            A dictionary mapping `XlColumn.xl_name` to the cell value for that
            row. Only columns present in both the table definition and the
            worksheet are included.

            Rows are skipped if:
            - A `primary` column is missing from the worksheet table, or
            - A `primary` column has an empty/falsy value, or
            - The resulting record dict would be empty.
        """
        if not table.ref:
            return

        min_col, min_row, max_col, max_row = range_boundaries(
            cast(str, table.ref)
        )
        assert (
            min_col is not None
            and min_row is not None
            and max_col is not None
            and max_row is not None
        )

        # Exclude Excel "Totals" rows from iteration when present.
        totals_row_count = table.totalsRowCount or 0
        if totals_row_count:
            max_row = max(min_row, max_row - totals_row_count)

        def _coerce_value(col: "XlColumn", value: Any) -> Any:
            """Coerce worksheet values based on column metadata.

            Excel cells may store numeric content as numbers even when the
            database column is textual. For columns declared as `string`, we
            convert simple numeric values to strings to avoid DB type errors and
            confusing diffs like `"10"` vs `10`.
            """
            if value is None:
                return None

            # Special handling for "unknown datetime" sentinel.
            if col.type_name in (
                "datetime",
                FIELD_TYPE_DT,
            ):
                if isinstance(value, str) and value.strip().lower() == "x":
                    return UNKNOWN_DATETIME
                if isinstance(value, datetime):
                    if (
                        value.year == 1000
                        and value.month == 2
                        and value.day == 3
                        and value.hour == 4
                        and value.minute == 5
                        and value.second == 6
                    ):
                        return UNKNOWN_DATETIME
                return value

            if getattr(col, "type_name", None) != "string":
                return value

            if isinstance(value, str):
                return value

            # Keep booleans stable (avoid "True"/"False" surprises).
            if isinstance(value, bool):
                return "1" if value else "0"

            # Convert numeric types. Prefer "10" over "10.0" when integral.
            if isinstance(value, int):
                return str(value)
            if isinstance(value, float):
                if value.is_integer():
                    return str(int(value))
                return str(value)

            try:
                from decimal import Decimal
            except ImportError:  # pragma: no cover
                Decimal = None  # type: ignore
            if Decimal is not None and isinstance(value, Decimal):
                if value == int(value):
                    return str(int(value))
                return str(value)

            return str(value)

        # Build map: table header name -> 0-based column index in worksheet.
        header_row = next(
            ws.iter_rows(
                min_row=min_row,
                max_row=min_row,
                min_col=min_col,
                max_col=max_col,
            )
        )
        col_name_to_idx = {
            str(cell.value): (self._get_ws_col_idx(cell) - min_col)
            for cell in header_row
        }

        for row in ws.iter_rows(
            min_row=min_row + 1,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        ):
            ignore_row = False
            xl_record = {}
            for c in self.columns:
                c_index = col_name_to_idx.get(c.xl_name, None)
                if c_index is None:
                    if c.primary:
                        # One of the primary columns is missing so
                        # we're not going to be able to work with the database.
                        ignore_row = True
                        break
                    continue
                value = row[c_index].value

                # Decide whether an empty primary key should make the row
                # unusable. For generated integer PKs (typically `id`),
                # allowing empty cells enables inserting new rows where the
                # database allocates the final id.
                if c.primary:
                    # Normalize "empty" values (None / whitespace-only string).
                    is_empty = value is None or (
                        isinstance(value, str) and value.strip() == ""
                    )

                    # Generated PK columns can be empty in Excel.
                    is_generated_pk = bool(getattr(c, "is_generated_pk", False))

                    if is_empty and not is_generated_pk:
                        # The primary key is incomplete.
                        ignore_row = True
                        break

                xl_record[c.xl_name] = _coerce_value(c, value)

            if ignore_row or not xl_record:
                continue

            yield xl_record

    def find_db_rec(
        self,
        session: "Session",
        xl_rec: Dict[str, Any],
        is_db_pk: Callable[[Any], bool],
    ) -> "T | None":
        """Locate the corresponding database record for an Excel row.

        Implementations typically use the table's primary columns to build a
        lookup query.

        Args:
            session: SQLAlchemy session used to query the database.
            xl_rec: Excel row dictionary mapping column names to values.
            is_db_pk: Predicate used to determine DB IDs. Defaults to
                `_default_is_db_pk`.

        Returns:
            The matching database record, or `None` if no match is found.
        """
        raise NotImplementedError

    def create_new_db_record(
        self, session: "Session", xl_rec: Dict[str, Any]
    ) -> T:
        """Create a new database record.

        Note that after this call each column in turn will get a chance to
        update this record.

        Args:
            session: SQLAlchemy session used for persistence and related
                lookups.
            xl_rec: Excel row dictionary mapping column names to values.

        Returns:
            The newly created database record instance.
        """
        raise NotImplementedError

    def apply_xl_to_db(
        self, session: "Session", db_rec: T, xl_rec: Dict[str, Any]
    ):
        """Apply an Excel row dictionary to a database record.

        This delegates to each column's `XlColumn.apply_xl_to_db()` with the
        corresponding value from `xl_rec` (or `None` if missing).

        Args:
            session: SQLAlchemy session used for lookups and persistence.
            db_rec: Database record to update.
            xl_rec: Excel row dictionary mapping column names to values.
        """
        # Derive `ua_*` values from the corresponding base column when both are
        # present. These fields are used for diacritics-insensitive searching
        # and should not be user-authored.
        try:
            from unidecode import unidecode  # type: ignore[import]
        except Exception:
            unidecode = None  # type: ignore

        if unidecode is not None:
            # Only compute when both ua_xxx and xxx exist in table definition.
            col_names = {c.xl_name for c in self.columns}
            for name in list(col_names):
                if not name.startswith("ua_"):
                    continue
                base = name[3:]
                if base not in col_names:
                    continue
                base_val = xl_rec.get(base, None)
                if base_val is None:
                    xl_rec[name] = None
                elif isinstance(base_val, str):
                    xl_rec[name] = unidecode(base_val)
                else:
                    xl_rec[name] = unidecode(str(base_val))

        for c in self.columns:
            if c.read_only:
                continue
            c.apply_xl_to_db(session, db_rec, xl_rec.get(c.xl_name, None))
