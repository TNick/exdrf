from typing import TYPE_CHECKING, Any, Dict, Tuple

from attrs import define, field

from exdrf_util.table2base import Range2Other

if TYPE_CHECKING:
    from openpyxl.worksheet.cell_range import CellRange  # type: ignore
    from reportlab.lib.styles import ParagraphStyle  # type: ignore
    from reportlab.platypus import Table  # type: ignore

FONT_SIZE_FACTOR = 0.8
FONT_NAME = "DejaVuSerifCondensed"
COLUMN_WIDTH_FACTOR = 2.54 * 2
IMAGE_SIZE_FACTOR = 0.8


def _initial_styles() -> list[Any]:
    """Initialize the styles for the pdf table.

    The generate method of the Range2Pdf will eventually create a TableStyle
    from this list of styles.
    """
    from reportlab.lib import colors  # type: ignore

    return [
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, -1), FONT_NAME),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
    ]


@define
class Range2Pdf(Range2Other):
    """Creates a pdf table from a range of cells using reportlab.

    Attributes:
        worksheet: (input) The worksheet to create the pdf table from.
        start_row: (input) The start row of the range.
        start_col: (input) The start column of the range.
        row_count: (input) The number of rows in the range.
        col_count: (input) The number of columns in the range.
        default_column_width: (input) The default width of a column in the
            pdf table. This is used when the column width is not set in the
            worksheet.
        crt_row: The current row number; used internally while creating the
            cells of the pdf table.
        crt_col: The current column number; used internally while creating the
            cells of the pdf table.
        table_data: The data for the pdf table. We accumulate the data for the
            pdf table here, row by row, cell by cell.
        row_data: The data for the current row, used while creating the cells
            of the pdf table.
        styles: The styles for the pdf table. We accumulate the styles for the
            pdf table here, then create a TableStyle from them.
        cache: The cache for lazy initialization of the paragraph styles.
        images: The images collected from the worksheet, indexed by cell key.
        str_cell_value: The value of the current cell. Various apply_...
            methods accumulate the value of the current cell here.
        par_style: The paragraph style for the current cell.
    """

    default_column_width: float = field(default=50, init=False)

    table_data: list[Any] = field(factory=list, init=False)
    row_data: list[Any] = field(factory=list, init=False)
    styles: list[Any] = field(factory=_initial_styles, init=False)
    cache: Dict[str, Any] = field(factory=dict, init=False)
    images: Dict[str, Tuple[str, float, float]] = field(
        factory=dict, init=False
    )
    str_cell_value: str = field(default="", init=False)
    par_style: "ParagraphStyle" = field(default=None, init=False)

    def generate(self) -> "Table":
        """Generate the pdf table."""
        # Collect individual cell data into rows and rows into table.
        for self.crt_row in range(self.start_row, self.last_row + 1):
            self.row_data = []
            for self.crt_col in range(self.start_col, self.last_col + 1):
                self.create_cell()
            self.table_data.append(self.row_data)

        # Merge cells.
        for merged in self.worksheet.merged_cells:
            self.merge_cells(merged)

        return self.create_table()

    def __attrs_post_init__(self):
        from openpyxl.utils import coordinate_to_tuple

        # Validate input parameters.
        self.validate_range()

        # Collect all images in the worksheet that are within the range.
        for image in self.worksheet._images:
            r, c = coordinate_to_tuple(image.anchor)
            if self.valid_coords(r, c):
                self.images[self.cell_key] = (
                    image.full_path,
                    image.width * IMAGE_SIZE_FACTOR,
                    image.height * IMAGE_SIZE_FACTOR,
                )

    @property
    def par_sty_center(self) -> "ParagraphStyle":
        """Get the paragraph style for center alignment."""
        value = self.cache.get("par_sty_center")
        if value is None:
            from reportlab.lib.enums import TA_CENTER  # type: ignore
            from reportlab.lib.styles import ParagraphStyle

            value = ParagraphStyle("Normal", None, alignment=TA_CENTER)
            self.cache["par_sty_center"] = value
        return value

    @property
    def par_sty_left(self) -> "ParagraphStyle":
        """Get the paragraph style for left alignment."""
        value = self.cache.get("par_sty_left")
        if value is None:
            from reportlab.lib.enums import TA_LEFT
            from reportlab.lib.styles import ParagraphStyle

            value = ParagraphStyle("Normal", None, alignment=TA_LEFT)
            self.cache["par_sty_left"] = value
        return value

    @property
    def par_sty_right(self) -> "ParagraphStyle":
        """Get the paragraph style for right alignment."""
        value = self.cache.get("par_sty_right")
        if value is None:
            from reportlab.lib.enums import TA_RIGHT
            from reportlab.lib.styles import ParagraphStyle

            value = ParagraphStyle("Normal", None, alignment=TA_RIGHT)
            self.cache["par_sty_right"] = value
        return value

    def create_cell(self):
        """Create the cell in the pdf table."""
        from reportlab.platypus import Paragraph  # type: ignore

        self.apply_borders()
        self.str_cell_value = str(self.xl_cell_value)

        self.apply_image()

        if not len(self.str_cell_value):
            self.row_data.append("")
            return

        xl_cell = self.xl_cell
        self.apply_font_face(
            FONT_NAME,  # cell.font.name
            xl_cell.font.sz,  # type: ignore
        )
        self.apply_font_color(xl_cell.font.color)  # type: ignore
        self.apply_bold(xl_cell.font.bold)  # type: ignore
        self.apply_italic(xl_cell.font.italic)  # type: ignore
        self.apply_underline(xl_cell.font.underline)  # type: ignore
        self.apply_font_background(
            xl_cell.fill.bgColor.rgb  # type: ignore
            if xl_cell.fill.bgColor
            else ""  # type: ignore
        )
        self.apply_h_align(xl_cell.alignment.horizontal)  # type: ignore
        self.apply_v_align(xl_cell.alignment.vertical)  # type: ignore
        self.row_data.append(Paragraph(self.str_cell_value, self.par_style))

    def apply_borders(self):
        """Creates inner grid and box styles for the current cell."""
        from reportlab.lib import colors  # type: ignore

        xl_cell = self.xl_cell
        if (
            xl_cell.border.left  # type: ignore
            and xl_cell.border.left.style is not None  # type: ignore
        ) or (
            xl_cell.border.top  # type: ignore
            and xl_cell.border.top.style is not None  # type: ignore
        ):
            self.styles.append(
                (
                    "INNERGRID",
                    (self.rel_col, self.rel_row),
                    (self.rel_col, self.rel_row),
                    0.25,
                    colors.black,
                )
            )
            self.styles.append(
                (
                    "BOX",
                    (self.rel_col, self.rel_row),
                    (self.rel_col, self.rel_row),
                    0.25,
                    colors.black,
                )
            )

    def apply_image(self) -> None:
        """Create the image in the pdf table."""
        img = self.images.get(self.cell_key)
        if not img:
            return

        self.str_cell_value = (
            f"{self.str_cell_value}<img "
            f'src="{img[0]}" '
            f'width="{img[1]}" '
            f'height="{img[2]}" valign="middle"/>'
        )

    def apply_font_face(self, font_name: str, size: int) -> None:
        """Apply the font to the value."""
        self.str_cell_value = (
            f'<font name="{font_name}" '
            f'size="{size * FONT_SIZE_FACTOR}">{self.str_cell_value}</font>'
        )

    def apply_font_color(self, color: str) -> None:
        """Apply the font color to the value."""
        if not color:
            return

        try:
            color = "0x{}".format(color[2:])
        except TypeError:
            color = "0x000000"

        self.styles.append(
            (
                "TEXTCOLOR",
                (self.rel_col, self.rel_row),
                (self.rel_col, self.rel_row),
                color,
            )
        )

    def apply_bold(self, is_set: bool = True) -> None:
        """Apply the bold to the value."""
        if not is_set:
            return

        self.str_cell_value = f"<b>{self.str_cell_value}</b>"

    def apply_italic(self, is_set: bool = True) -> None:
        """Apply the italic to the value."""
        if not is_set:
            return

        self.str_cell_value = f"<i>{self.str_cell_value}</i>"

    def apply_underline(self, is_set: bool = True) -> None:
        """Apply the underline to the value."""
        if not is_set:
            return

        self.str_cell_value = f"<u>{self.str_cell_value}</u>"

    def apply_font_background(self, color: str) -> None:
        """Apply the font background to the value."""
        if not color:
            return

        if color != "00000000":
            self.styles.append(
                (
                    "BACKGROUND",
                    (self.rel_col, self.rel_row),
                    (self.rel_col, self.rel_row),
                    "0x{}".format(color[2:]),
                )
            )

    def apply_h_align(self, horizontal: str) -> None:
        if not horizontal:
            self.par_style = self.par_sty_center
            return
        if horizontal == "left":
            self.par_style = self.par_sty_left
        elif horizontal == "right":
            self.par_style = self.par_sty_right
        else:
            self.par_style = self.par_sty_center
        self.styles.append(
            (
                "ALIGN",
                (self.rel_col, self.rel_row),
                (self.rel_col, self.rel_row),
                horizontal.upper(),
            )
        )

    def apply_v_align(self, vertical: str) -> None:
        if not vertical:
            return
        self.styles.append(
            (
                "VALIGN",
                (self.rel_col, self.rel_row),
                (self.rel_col, self.rel_row),
                vertical.upper(),
            )
        )

    @property
    def pdf_column_widths(self) -> list[float]:
        """Get the width of each column in the pdf table."""
        from openpyxl.utils import get_column_letter

        col_w = []
        for c in range(self.start_col, self.last_col + 1):
            xl_width = self.worksheet.column_dimensions[
                get_column_letter(c)
            ].width
            cw = xl_width * COLUMN_WIDTH_FACTOR
            if not cw:
                cw = self.default_column_width
            col_w.append(cw)
        return col_w

    def merge_cells(self, merged: "CellRange") -> None:
        """Merge the cells in the pdf table.

        Args:
            merged: The merged cells to merge.
        """
        if merged.min_row > self.last_row or merged.max_row > self.last_row:
            return
        if merged.min_row < self.start_row or merged.max_row < self.start_row:
            return
        if merged.min_col > self.last_col or merged.max_col > self.last_col:
            return
        if merged.min_col < self.start_col or merged.max_col < self.start_col:
            return

        # Collect values from across the merged cells.
        values = []
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                cell = self.table_data[r - self.start_row][c - self.start_col]
                if cell != "":
                    values.append(cell)
                    self.table_data[r - self.start_row][c - self.start_col] = ""
        if len(values) == 0:
            value = ""
        elif len(values) == 1:
            value = values[0]
        else:
            value = ", ".join(values)

        # Replace the merged cells with the combined value.
        delta_row = merged.min_row - self.start_row
        delta_col = merged.min_col - self.start_col
        self.table_data[delta_row][delta_col] = value
        self.styles.append(
            (
                "SPAN",
                (delta_col, delta_row),
                (
                    merged.max_col - self.start_col,
                    merged.max_row - self.start_row,
                ),
            )
        )

    def create_table(self) -> "Table":
        """Create the pdf table."""
        from reportlab.platypus import Table, TableStyle  # type: ignore

        table = Table(self.table_data, colWidths=self.pdf_column_widths)
        table.setStyle(TableStyle(self.styles))

        return table
