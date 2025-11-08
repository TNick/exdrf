import logging
from typing import TYPE_CHECKING, Any, Optional, cast

from attrs import define, field
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.cell import Cell, MergedCell
    from openpyxl.cell.read_only import ReadOnlyCell  # type: ignore
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        NamedStyle,
        PatternFill,
        Side,
    )
    from openpyxl.worksheet.worksheet import Worksheet


logger = logging.getLogger(__name__)


@define
class XlCreatorBase:
    """Base class for writing tables.

    For cells merged across columns the Excel seems unable to properly
    calculate the height of the merged cells. To alleviate that we use the
    following trick: when the user requests the h_trick in set_cell, we
    use the next available column, resize it to the width of the merged cells
    and set the text there. We then hide that column and Excel will use the
    proper height. Note that this does not work for cells that are also
    merged across rows.

    Attributes:
        workbook: The openpyxl workbook to write to.
        worksheet: The openpyxl worksheet to write to. Defaults to the
            active worksheet in the workbook.
        column_widths: The widths of the columns. Defaults to an empty list.
            The number of entries in this list indicates the number of columns.
        empty_character: The character to use for empty cells. Defaults to "-".
        crt_row: The current row index. Starts at 1.
        h_trick_counter: The counter for the h-trick. Starts at 1.
    """

    workbook: "Workbook"
    worksheet: "Worksheet" = field(default=None)
    empty_character: str = field(default="-")
    column_widths: list[float] = field(default=[])
    crt_row: int = field(default=1, init=False)
    h_trick_counter: int = field(default=1, init=False)
    style_cache: dict[str, Any] = field(factory=dict, init=False)

    def __attrs_post_init__(self) -> None:
        """Provide dynamic defaults."""
        self.worksheet = self.workbook.active

    def setup_page(self, portrait: bool = False):
        self.worksheet.print_options.horizontalCentered = True
        self.worksheet.print_options.verticalCentered = False
        # Fit one page wide, automatic height
        self.worksheet.page_setup.fitToWidth = 1
        self.worksheet.page_setup.fitToHeight = 0

        # Ensure fit-to-page is used instead of explicit scaling
        self.worksheet.page_setup.scale = None  # type: ignore[assignment]

        # Enable fit-to-page on sheet properties when supported
        from openpyxl.worksheet.properties import PageSetupProperties

        sheet_props = self.worksheet.sheet_properties
        page_setup_pr = getattr(sheet_props, "pageSetUpPr", None)
        if page_setup_pr is None:
            sheet_props.pageSetUpPr = PageSetupProperties(fitToPage=True)
        else:
            page_setup_pr.fitToPage = True

        self.worksheet.page_margins.left = 0.4 if portrait else 0.2
        self.worksheet.page_margins.right = 0.2
        self.worksheet.page_margins.top = 0.2 if portrait else 0.4
        self.worksheet.page_margins.bottom = 0.2
        self.worksheet.page_margins.header = 0.0
        self.worksheet.page_margins.footer = 0.0

    def setup_a4_landscape_page(self):
        """Setup the worksheet for an A4 landscape page."""
        self.worksheet.page_setup.orientation = (
            self.worksheet.ORIENTATION_LANDSCAPE
        )  # type: ignore
        self.worksheet.page_setup.paperSize = (
            self.worksheet.PAPERSIZE_A4
        )  # type: ignore
        self.setup_page(False)

    def setup_a4_portrait_page(self):
        """Setup the worksheet for an A4 portrait page."""
        self.worksheet.page_setup.orientation = (
            self.worksheet.ORIENTATION_PORTRAIT
        )  # type: ignore
        self.worksheet.page_setup.paperSize = (
            self.worksheet.PAPERSIZE_A4
        )  # type: ignore
        self.setup_page(True)

    def setup_a3_landscape_page(self):
        """Setup the worksheet for an A3 landscape page."""
        self.worksheet.page_setup.orientation = (
            self.worksheet.ORIENTATION_LANDSCAPE
        )  # type: ignore
        self.worksheet.page_setup.paperSize = (
            self.worksheet.PAPERSIZE_A3
        )  # type: ignore
        self.setup_page(False)

    def setup_a3_portrait_page(self):
        """Setup the worksheet for an A3 portrait page."""
        self.worksheet.page_setup.orientation = (
            self.worksheet.ORIENTATION_PORTRAIT
        )  # type: ignore
        self.worksheet.page_setup.paperSize = (
            self.worksheet.PAPERSIZE_A3
        )  # type: ignore
        self.setup_page(True)

    @property
    def grey_fill(self) -> "PatternFill":
        """Get the solid grey fill style."""
        value = self.style_cache.get("grey_fill")
        if value is None:
            from openpyxl.styles import PatternFill

            value = PatternFill(
                fill_type="solid", start_color="FFEEEEEE", end_color="FFEEEEEE"
            )
            self.style_cache["grey_fill"] = value
        return value

    @property
    def side_thick(self) -> "Side":
        """Get the thick side style."""
        value = self.style_cache.get("side_thick")
        if value is None:
            from openpyxl.styles import Side

            value = Side(style="thick", color="000000")
            self.style_cache["side_thick"] = value
        return value

    @property
    def side_thin(self) -> "Side":
        """Get the thin side style."""
        value = self.style_cache.get("side_thin")
        if value is None:
            from openpyxl.styles import Side

            value = Side(style="thin", color="000000")
            self.style_cache["side_thin"] = value
        return value

    @property
    def font_12_b(self) -> "Font":
        """Get the 12pt bold font."""
        value = self.style_cache.get("font_12_b")
        if value is None:
            from openpyxl.styles import Font

            value = Font(bold=True, size=12)
            self.style_cache["font_12_b"] = value
        return value

    @property
    def font_12(self) -> "Font":
        """Get the 12pt font."""
        value = self.style_cache.get("font_12")
        if value is None:
            from openpyxl.styles import Font

            value = Font(bold=False, size=12)
            self.style_cache["font_12"] = value
        return value

    @property
    def font_11(self) -> "Font":
        """Get the 11pt font."""
        value = self.style_cache.get("font_11")
        if value is None:
            from openpyxl.styles import Font

            value = Font(bold=False, size=11)
            self.style_cache["font_11"] = value
        return value

    @property
    def font_11_b(self) -> "Font":
        """Get the 11pt bold font."""
        value = self.style_cache.get("font_11_b")
        if value is None:
            from openpyxl.styles import Font

            value = Font(bold=True, size=11)
            self.style_cache["font_11_b"] = value
        return value

    @property
    def font_10(self) -> "Font":
        """Get the 10pt font."""
        value = self.style_cache.get("font_10")
        if value is None:
            from openpyxl.styles import Font

            value = Font(bold=False, size=10)
            self.style_cache["font_10"] = value
        return value

    @property
    def font_10_b(self) -> "Font":
        """Get the 10pt bold font."""
        value = self.style_cache.get("font_10_b")
        if value is None:
            from openpyxl.styles import Font

            value = Font(bold=True, size=10)
            self.style_cache["font_10_b"] = value
        return value

    @property
    def font_09(self) -> "Font":
        """Get the 9pt font."""
        value = self.style_cache.get("font_09")
        if value is None:
            from openpyxl.styles import Font

            value = Font(bold=False, size=9)
            self.style_cache["font_09"] = value
        return value

    @property
    def font_08(self) -> "Font":
        """Get the 8pt font."""
        value = self.style_cache.get("font_08")
        if value is None:
            from openpyxl.styles import Font

            value = Font(bold=False, size=8)
            self.style_cache["font_08"] = value
        return value

    @property
    def font_08_b(self) -> "Font":
        """Get the 8pt bold font."""
        value = self.style_cache.get("font_08_b")
        if value is None:
            from openpyxl.styles import Font

            value = Font(bold=True, size=8)
            self.style_cache["font_08_b"] = value
        return value

    @property
    def border_thin(self) -> "Border":
        """Get the thin border style."""
        value = self.style_cache.get("border_thin")
        if value is None:
            from openpyxl.styles import Border

            side_thin = self.side_thin
            value = Border(
                left=side_thin,
                top=side_thin,
                right=side_thin,
                bottom=side_thin,
                vertical=side_thin,  # type: ignore
                horizontal=side_thin,  # type: ignore
            )
            self.style_cache["border_thin"] = value
        return value

    @property
    def border_thick(self) -> "Border":
        """Get the thick border style."""
        value = self.style_cache.get("border_thick")
        if value is None:
            from openpyxl.styles import Border

            side_thick = self.side_thick
            value = Border(
                left=side_thick,
                top=side_thick,
                right=side_thick,
                bottom=side_thick,
                vertical=side_thick,  # type: ignore
                horizontal=side_thick,  # type: ignore
            )
            self.style_cache["border_thick"] = value
        return value

    @property
    def align_center_vcenter(self) -> "Alignment":
        """Get the center alignment style."""
        value = self.style_cache.get("align_center_vcenter")
        if value is None:
            from openpyxl.styles import Alignment

            value = Alignment(
                horizontal="center",
                vertical="center",
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0,
            )
            self.style_cache["align_center_vcenter"] = value
        return value

    @property
    def align_left_vcenter(self) -> "Alignment":
        """Get the center alignment style."""
        value = self.style_cache.get("align_left_vcenter")
        if value is None:
            from openpyxl.styles import Alignment

            value = Alignment(
                horizontal="left",
                vertical="center",
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0,
            )
            self.style_cache["align_left_vcenter"] = value
        return value

    @property
    def align_right_vcenter(self) -> "Alignment":
        """Get the center alignment style."""
        value = self.style_cache.get("align_right_vcenter")
        if value is None:
            from openpyxl.styles import Alignment

            value = Alignment(
                horizontal="right",
                vertical="center",
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0,
            )
            self.style_cache["align_right_vcenter"] = value
        return value

    @property
    def align_center_vtop(self) -> "Alignment":
        """Get the center alignment style."""
        value = self.style_cache.get("align_center_vtop")
        if value is None:
            from openpyxl.styles import Alignment

            value = Alignment(
                horizontal="center",
                vertical="top",
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0,
            )
            self.style_cache["align_center_vtop"] = value
        return value

    @property
    def align_left_vtop(self) -> "Alignment":
        """Get the center alignment style."""
        value = self.style_cache.get("align_left_vtop")
        if value is None:
            from openpyxl.styles import Alignment

            value = Alignment(
                horizontal="left",
                vertical="top",
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0,
            )
            self.style_cache["align_left_vtop"] = value
        return value

    @property
    def align_right_vtop(self) -> "Alignment":
        """Get the center alignment style."""
        value = self.style_cache.get("align_right_vtop")
        if value is None:
            from openpyxl.styles import Alignment

            value = Alignment(
                horizontal="right",
                vertical="top",
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0,
            )
            self.style_cache["align_right_vtop"] = value
        return value

    @property
    def align_center_vbtm(self) -> "Alignment":
        """Get the center alignment style."""
        value = self.style_cache.get("align_center_vbtm")
        if value is None:
            from openpyxl.styles import Alignment

            value = Alignment(
                horizontal="center",
                vertical="bottom",
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0,
            )
            self.style_cache["align_center_vbtm"] = value
        return value

    @property
    def align_left_vbtm(self) -> "Alignment":
        """Get the center alignment style."""
        value = self.style_cache.get("align_left_vbtm")
        if value is None:
            from openpyxl.styles import Alignment

            value = Alignment(
                horizontal="left",
                vertical="bottom",
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0,
            )
            self.style_cache["align_left_vbtm"] = value
        return value

    @property
    def align_right_vbtm(self) -> "Alignment":
        """Get the center alignment style."""
        value = self.style_cache.get("align_right_vbtm")
        if value is None:
            from openpyxl.styles import Alignment

            value = Alignment(
                horizontal="right",
                vertical="bottom",
                text_rotation=0,
                wrap_text=True,
                shrink_to_fit=False,
                indent=0,
            )
            self.style_cache["align_right_vbtm"] = value
        return value

    def cell(self, row: int, col: int) -> "MergedCell | ReadOnlyCell | Cell":
        """Get the cell at the given row and column.

        When the row is negative, the sign is discarded and the value is
        added to the current row, so we can only address the current row and
        rows after that using this mechanism.

        Examples, assuming the crt_row is 10 and the column widths are
        [10, 20, 10]:
            - row = 1, col = 1, value = "Hello" -> sets the cell A1
            - row = 0, col = 1, value = "Hello" -> sets the cell A10
            - row = -1, col = 2, value = "Hello" -> sets the cell B11
            - row = 1, col = 0, value = "Hello" -> sets the cell C1
            - row = 1, col = -1, value = "Hello" -> sets the cell B1

        Args:
            row: The row index. This can be an absolute row (when this is a
                positive integer) or a relative row (when this is 0 or a
                negative integer).
            col: The column index. This can be an absolute column (when this is
                a positive integer) or relative to the last column (when this is
                a negative integer).
        """
        row = row if row > 0 else self.crt_row - row
        if col <= 0:
            assert len(self.column_widths) > 0, "Column widths must be set"
            col = len(self.column_widths) - col

        return self.worksheet.cell(row=row, column=col)

    def set_cell(
        self,
        row: int,
        col: int,
        value: Any,
        merge_rows: int = 1,
        merge_cols: int = 1,
        style: "Optional[str | NamedStyle]" = None,
        empty_character: Optional[str] = None,
        h_trick: bool = False,
    ) -> "Cell | MergedCell":
        """Set the value of a cell.

        When the row is negative, the sign is discarded and the value is
        added to the current row, so we can only address the current row and
        rows after that using this mechanism.

        Examples, assuming the crt_row is 10 and the column widths are
        [10, 20, 10]:
            - row = 1, col = 1, value = "Hello" -> sets the cell A1
            - row = 0, col = 1, value = "Hello" -> sets the cell A10
            - row = -1, col = 2, value = "Hello" -> sets the cell B11
            - row = 1, col = 0, value = "Hello" -> sets the cell C1
            - row = 1, col = -1, value = "Hello" -> sets the cell B1

        Args:
            row: The row index. This can be an absolute row (when this is a
                positive integer) or a relative row (when this is 0 or a
                negative integer).
            col: The column index. This can be an absolute column (when this is
                a positive integer) or relative to the last column (when this is
                a negative integer).
            value: The value to set.
            merge_rows: The number of rows to merge. Defaults to 1, which means
                the cell will not be merged.
            merge_cols: The number of columns to merge. Defaults to 1, which
                means the cell will not be merged.
            style: The optional style to apply to the cell.
            empty_character: Override the default character to use for
                empty cells.
            h_trick: Whether to use the h-trick to compute the height of the
                cell. This is useful when the cells are merged and excel
                will not compute the height of the cell correctly.

        Returns:
            The cell that was set.
        """

        # Compute an absolute row and column index.
        row = row if row > 0 else self.crt_row - row
        if col <= 0:
            assert len(self.column_widths) > 0, "Column widths must be set"
            col = len(self.column_widths) - col

        # Get this cell. We assume this is not a read-only cell.
        cell = cast(
            "Cell | MergedCell", self.worksheet.cell(row=row, column=col)
        )
        assert cell.__class__.__name__ in ["Cell", "MergedCell"]

        # Replace with empty character if the value is None or empty.
        if value is None or value == "":
            value = (
                empty_character
                if empty_character is not None
                else self.empty_character
            )

        try:
            # Set the value.
            cell.value = value

            # Merge the cells if requested.
            if merge_rows > 1 or merge_cols > 1:
                self.worksheet.merge_cells(
                    start_row=row,
                    start_column=col,
                    end_row=row + merge_rows - 1,
                    end_column=col + merge_cols - 1,
                )
                if style:
                    self.set_range_style(
                        row=row,
                        col=col,
                        merge_rows=merge_rows,
                        merge_cols=merge_cols,
                        style=style,
                    )
            elif style:
                cell.style = (
                    style
                    if isinstance(style, str)
                    else style.name  # type: ignore
                )

            # Apply the h-trick if requested.
            if h_trick and value:
                self.apply_h_trick(
                    row=row,
                    col=col,
                    merge_cols=merge_cols,
                    value=value,
                    style=style,
                )
        except Exception:
            logger.exception(
                "Failed to set the value of cell at row %s and column %s to %s",
                row,
                col,
                value,
            )

        return cell

    def set_range_empty(
        self,
        row: int,
        col: int,
        row_count: int,
        col_count: int,
        style: "Optional[str | NamedStyle]" = None,
        empty_character: Optional[str] = None,
    ):
        """Set the value of a range of cells to the empty character.

        Args:
            row: The row index.
            col: The column index.
            row_count: The number of rows to set.
            col_count: The number of columns to set.
            style: The optional style to apply to the cells.
            empty_character: Override the default character to use for
                empty cells.
        """
        if empty_character is None:
            empty_character = self.empty_character

        for r in range(row, row + row_count):
            for c in range(col, col + col_count):
                self.set_cell(
                    row=r,
                    col=c,
                    value=empty_character,
                    style=style,
                )

    def set_range_style(
        self,
        row: int,
        col: int,
        merge_rows: int,
        merge_cols: int,
        style: "str | NamedStyle",
    ) -> None:
        """Set the style to a range of cells.

        Args:
            row: The row index.
            col: The column index.
            merge_rows: The number of rows to merge.
            merge_cols: The number of columns to merge.
            style: The style to apply to the cells.
        """
        style_name = (
            style if isinstance(style, str) else style.name  # type: ignore
        )
        for rs in range(0, merge_rows):
            for cs in range(0, merge_cols):
                crt_cell = cast(
                    "Cell | MergedCell",
                    self.worksheet.cell(row=row + rs, column=col + cs),
                )
                assert crt_cell.__class__.__name__ in ["Cell", "MergedCell"]
                crt_cell.style = style_name

    def apply_h_trick(
        self,
        row: int,
        col: int,
        merge_cols: int,
        value: Any,
        style: "Optional[str | NamedStyle]" = None,
    ) -> "Cell | MergedCell":
        """Apply the h-trick to a cell.

        Args:
            row: The row index.
            col: The column index.
            merge_cols: The number of columns to merge.
            value: The value to set.
            style: The optional style to apply to the cell.
        """
        self.h_trick_counter += 1

        ghost_col = col + len(self.column_widths) + self.h_trick_counter

        ghost = cast(
            "Cell | MergedCell",
            self.worksheet.cell(
                row=row,
                column=ghost_col,
            ),
        )
        assert ghost.__class__.__name__ in ["Cell", "MergedCell"]

        # Retrieve the dimension object for the ghost column.
        cd = self.worksheet.column_dimensions[get_column_letter(ghost_col)]

        # Set its width to the width of the merged range of columns.
        cd.width = self.merged_width(col, merge_cols)
        cd.auto_size = False
        cd.hidden = True

        # Use same style as the original cell.
        if style:
            ghost.style = (
                style if isinstance(style, str) else style.name  # type: ignore
            )

        # Set the value of the ghost cell.
        ghost.value = value

        return ghost

    def merged_width(self, col: int, count: int) -> float:
        """Compute the width of a merged range of columns.

        Args:
            col: The column index.
            count: The number of columns to merge.
        """
        assert col > 0, "Column index must be positive"
        assert count > 0, "Number of columns to merge must be positive"
        assert col + count - 1 <= len(
            self.column_widths
        ), "Column index and number of columns to merge must be compatible"
        return sum(self.column_widths[col - 1 : col + count - 1])

    def sanitize_sheet_title(self, value: str) -> str:
        # Excel sheet titles cannot contain: : \ / ? * [ ] and must be <= 31
        for ch in [":", "\\", "/", "?", "*", "[", "]"]:
            value = value.replace(ch, "-")
        value = value.strip()
        if len(value) > 31:
            value = value[:31]
        return value or "Sheet"

    def generate(self):
        """Generates content into the worksheet.

        This is the method that subclasses need to implement.
        """
        raise NotImplementedError
